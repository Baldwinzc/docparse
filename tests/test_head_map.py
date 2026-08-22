from __future__ import annotations

import io
from pathlib import Path

import pytest

pytest.importorskip("openpyxl")

from openpyxl import Workbook
from openpyxl.styles import Border, Side

from docparse.adapters.parsers.excel import parse_excel
from docparse.extraction.head_map import map_document_head, map_sheet_head
from docparse.schema.loader import load_schema

_DEMO = Path("/Users/baldwin/Desktop/taizhou/AI识别Demo")
REAL_HENGXIN = _DEMO / "（恒信）一般贸易草单HDX260251BLU.xlsx"
REAL_GUOGUANG = _DEMO / "（国光）箱单发票合同26VN0502-1.xlsx"


def _thin() -> Border:
    line = Side(style="thin")
    return Border(left=line, right=line, top=line, bottom=line)


def _workbook(builders: dict) -> bytes:
    book = Workbook()
    first = True
    for title, fill in builders.items():
        sheet = book.active if first else book.create_sheet(title)
        if first:
            sheet.title = title
            first = False
        fill(sheet)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _draft(sheet) -> None:
    sheet["A1"] = "中华人民共和国海关出口货物报关单"
    sheet.merge_cells("A3:D3")
    sheet["A3"] = "境内发货人"
    sheet.merge_cells("A4:D4")
    sheet["A4"] = "惠州市恒德信精密科技有限公司441394164D"
    sheet["E3"] = "出境关别"
    sheet["E4"] = "莲塘口岸"
    sheet.merge_cells("A5:D5")
    sheet["A5"] = "境外收货人"
    sheet.merge_cells("A6:D6")
    sheet["A6"] = "BLU PRECISION LIMITED"
    sheet.merge_cells("A7:D7")
    sheet["A7"] = "生产销售单位"
    sheet.merge_cells("A8:D8")
    sheet["A8"] = "惠州市恒德信精密科技有限公司"
    sheet.merge_cells("A9:D9")
    sheet["A9"] = "合同协议号"
    sheet.merge_cells("A10:D10")
    sheet["A10"] = "HDX2026-251"
    sheet.merge_cells("A11:C11")
    sheet["A11"] = "包装种类"
    sheet["D11"] = "件数"
    sheet["E11"] = "毛重（千克）"
    sheet["F11"] = "净重（千克）"
    sheet.merge_cells("G11:H11")
    sheet["G11"] = "成交方式"
    sheet.merge_cells("A12:C12")
    sheet["A12"] = "其它"
    sheet["D12"] = 40
    sheet["E12"] = 296.46
    sheet["F12"] = 218.375
    sheet.merge_cells("G12:H12")
    sheet["G12"] = "FOB"
    sheet["I11"] = "运费"
    sheet["I12"] = "3 USD"
    sheet["A17"] = "项号"
    sheet["B17"] = "商品编号"
    sheet["C17"] = "商品名称及规格型号"
    sheet["D17"] = "申报要素"
    sheet["A18"] = 1
    sheet["B18"] = "9111900000"
    for row in sheet.iter_rows(min_row=1, max_row=18, min_col=1, max_col=9):
        for cell in row:
            cell.border = _thin()


def _guoguang_packing(sheet) -> None:
    sheet["E5"] = "装箱单 PACKING LIST"
    sheet["A6"] = "日期DATE:"
    sheet["B6"] = "2026-03-15"
    sheet["A7"] = "发票/INVOICE NO.:"
    sheet["B7"] = "26VN0502-1"
    sheet["G7"] = "合同号CONTRACT NO.:"
    sheet["H7"] = "26VN0502"
    sheet["A8"] = "卖方SELLER"
    sheet["A9"] = "GUOGUANG ELECTRIC COMPANY LIMITED."
    sheet["F8"] = "买方BUYER"
    sheet["F9"] = "GUOGUANG ACOUSTICS (VIETNAM) COMPANY LIMITED"
    sheet["B12"] = "物料名称"
    sheet["D12"] = "出货数量"
    sheet["F12"] = "总箱数"
    sheet["B13"] = "贴纸"
    sheet["D13"] = 100
    for row in sheet.iter_rows(min_row=5, max_row=13, min_col=1, max_col=10):
        for cell in row:
            cell.border = _thin()


def _auxiliary_with_contract(sheet) -> None:
    sheet["B1"] = "报关单号"
    sheet["C1"] = "报关日期"
    sheet["K1"] = "SAP编号"
    sheet["A3"] = "合同协议号"
    sheet["A4"] = "SHOULD-NOT-MAP"
    for row in sheet.iter_rows(min_row=1, max_row=4, min_col=1, max_col=11):
        for cell in row:
            cell.border = _thin()


def _by_name(fields) -> dict[str, object]:
    grouped: dict[str, list] = {}
    for field in fields:
        grouped.setdefault(field.name, []).append(field)
    return grouped


def test_schema_head_map_flags() -> None:
    schema = load_schema()
    assert schema.field("tradeName").head_map == "trailing_code"
    assert schema.field("tradeName").split_target == "tradeCode"
    assert schema.field("tradeCode").head_map == "skip"
    assert schema.field("ownerName").head_map == "trailing_code"
    assert schema.field("feeMark").head_map == "skip"
    assert schema.field("cusVoyageNo").head_map == "skip"
    assert schema.field("agentCode").parse is False
    assert "卖方SELLER" in schema.field("tradeName").anchors
    assert "Bill To" in schema.field("consignorEname").anchors
    assert "合同号CONTRACT NO." in schema.field("contrNo").anchors


def test_hengxin_draft_head_fields() -> None:
    document = parse_excel(
        _workbook({"一般贸易出口": _draft}),
        file_id="hx",
        filename="hengxin-draft.xlsx",
    )
    draft = document.sheets[0]
    assert draft.role == "draft"
    fields = map_sheet_head(draft, document)
    by_name = {item.name: item for item in fields}

    assert by_name["contrNo"].value == "HDX2026-251"
    assert by_name["consignorEname"].value == "BLU PRECISION LIMITED"
    assert by_name["packNo"].value == "40"
    assert by_name["grossWt"].value == "296.46"
    assert by_name["netWt"].value == "218.375"
    assert by_name["transMode"].value == "FOB"
    assert by_name["tradeName"].value == "惠州市恒德信精密科技有限公司"
    assert by_name["tradeCode"].value == "441394164D"
    assert by_name["ownerName"].value == "惠州市恒德信精密科技有限公司"
    assert "ownerCode" not in by_name
    assert "feeMark" not in by_name
    assert "feeRate" not in by_name
    assert "agentCode" not in by_name
    assert "agentName" not in by_name

    for name in (
        "contrNo",
        "consignorEname",
        "packNo",
        "grossWt",
        "netWt",
        "transMode",
        "tradeName",
        "tradeCode",
    ):
        field = by_name[name]
        assert field.status.value == "accepted"
        assert field.evidence
        assert field.evidence[0].cell.startswith("一般贸易出口!")
        assert field.evidence[0].quote.startswith("一般贸易出口!")


def test_guoguang_packing_contract_and_parties() -> None:
    document = parse_excel(
        _workbook({"总箱单": _guoguang_packing}),
        file_id="gg",
        filename="guoguang-packing.xlsx",
    )
    packing = document.sheets[0]
    assert packing.role == "packing"
    fields = map_sheet_head(packing, document)
    by_name = {item.name: item for item in fields}

    assert by_name["contrNo"].value == "26VN0502"
    assert by_name["tradeName"].value == "GUOGUANG ELECTRIC COMPANY LIMITED."
    assert by_name["consignorEname"].value == "GUOGUANG ACOUSTICS (VIETNAM) COMPANY LIMITED"
    assert all(item.evidence for item in fields)
    assert "billNo" not in by_name
    assert not any(item.value == "26VN0502-1" for item in fields)


def test_exclude_sheet_is_not_mapped() -> None:
    document = parse_excel(
        _workbook({"Sheet3": _auxiliary_with_contract}),
        file_id="aux",
        filename="aux.xlsx",
    )
    sheet = document.sheets[0]
    assert sheet.consume == "exclude"
    assert map_sheet_head(sheet, document) == []
    assert map_document_head(document) == []


def test_multi_sheet_keeps_separate_candidates() -> None:
    document = parse_excel(
        _workbook({"一般贸易出口": _draft, "总箱单": _guoguang_packing}),
        file_id="mix",
        filename="mix.xlsx",
    )
    fields = map_document_head(document)
    grouped = _by_name(fields)
    contracts = {item.value for item in grouped["contrNo"]}
    assert contracts == {"HDX2026-251", "26VN0502"}
    cells = {item.evidence[0].cell for item in grouped["contrNo"]}
    assert any(cell.startswith("一般贸易出口!") for cell in cells)
    assert any(cell.startswith("总箱单!") for cell in cells)


@pytest.mark.skipif(not REAL_HENGXIN.exists(), reason="本地恒信样本不在 CI")
def test_hengxin_sample_head() -> None:
    document = parse_excel(
        REAL_HENGXIN.read_bytes(),
        file_id="hx-real",
        filename=REAL_HENGXIN.name,
    )
    draft = next(sheet for sheet in document.sheets if sheet.role == "draft")
    by_name = {item.name: item for item in map_sheet_head(draft, document)}
    assert by_name["contrNo"].value == "HDX2026-251"
    assert by_name["consignorEname"].value == "BLU PRECISION LIMITED"
    assert by_name["packNo"].value == "40"
    assert by_name["grossWt"].value == "296.46"
    assert by_name["netWt"].value == "218.375"
    assert by_name["transMode"].value == "FOB"
    assert by_name["tradeCode"].value == "441394164D"
    assert all(item.evidence for item in by_name.values())


@pytest.mark.skipif(not REAL_GUOGUANG.exists(), reason="本地国光样本不在 CI")
def test_guoguang_sample_head() -> None:
    document = parse_excel(
        REAL_GUOGUANG.read_bytes(),
        file_id="gg-real",
        filename=REAL_GUOGUANG.name,
    )
    packing = next(sheet for sheet in document.sheets if sheet.role == "packing")
    by_name = {item.name: item for item in map_sheet_head(packing, document)}
    assert by_name["contrNo"].value == "26VN0502"
    assert by_name["tradeName"].value == "GUOGUANG ELECTRIC COMPANY LIMITED."
    assert by_name["consignorEname"].value == "GUOGUANG ACOUSTICS (VIETNAM) COMPANY LIMITED"
    assert not any(item.value == "26VN0502-1" for item in by_name.values())
    assert all(item.evidence for item in by_name.values())
