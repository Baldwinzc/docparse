"""锚点归一化与别名补齐（#66）：键形状、纯代码值路由、列映射优先级。"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest

pytest.importorskip("openpyxl")

from openpyxl import Workbook

from docparse.adapters.parsers.excel import parse_excel
from docparse.extraction.assemble import assemble_declaration
from docparse.extraction.goods_map import map_sheet_goods
from docparse.extraction.head_map import map_sheet_head
from docparse.schema.loader import load_code_tables, load_layout_vocab, load_schema
from docparse.schema.textnorm import fold_key, fold_spaced, strip_trailing_code

_SAMPLES = Path("/Users/baldwin/Desktop/taizhou/补充测试")
REAL_GSC = _SAMPLES / "GSC出仓QPGSCO260617006A.xlsx"
REAL_GSRUA = _SAMPLES / "GSRUA26601CLLG01(1).xlsx"
REAL_TONDA2 = _SAMPLES / "通达2.xlsx"
REAL_DUOKE = _SAMPLES / "6-17 多科报关资料香港出货 DKTX-2606057 多科通讯乐乐高(1).xls"
REAL_DONNELLEY = _SAMPLES / "202606 R26JU551-Y报关一般.xls"


def _workbook(builders: dict) -> bytes:
    book = Workbook()
    first = True
    for title, fill in builders.items():
        sheet = book.active if first else book.create_sheet(title)
        if first:
            sheet.title = title
            first = False
        fill(sheet)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _by_name(fields) -> dict:
    return {item.name: item for item in fields}


# ---- 键归一化 ----


def test_fold_key_strips_all_whitespace_and_fullwidth_parens() -> None:
    assert fold_key("毛 重") == fold_key("毛重")
    assert fold_key("毛重\n（公斤）") == fold_key("毛重(公斤)")
    assert fold_key("淨    重") == "淨重"


def test_fold_key_strips_trailing_alnum_code_not_words() -> None:
    assert fold_key("贸易方式（0110）") == fold_key("贸易方式")
    assert fold_key("征免性质(101)") == fold_key("征免性质")
    # 括号里不是字母数字的不剥：这是语义的一部分
    assert fold_key("贸易国（地区）") != fold_key("贸易国")
    assert fold_key("毛重（公斤）") != fold_key("毛重")


def test_strip_trailing_code_variants() -> None:
    assert strip_trailing_code("申报要素（A1B2）") == "申报要素"
    assert strip_trailing_code("监管方式") == "监管方式"


def test_fold_spaced_keeps_token_boundaries() -> None:
    assert fold_spaced("Unit Price") == "unit price"
    assert fold_spaced("毛重\n（公斤）") == "毛重(公斤)"


# ---- 尾码标签：版面刀认出 + head_map 命中 ----


def _duoke_label_sheet(sheet) -> None:
    sheet["A1"] = "中华人民共和国海关出口货物报关单"
    sheet["A3"] = "境内发货人"
    sheet["A4"] = "深圳市多科通讯有限公司"
    sheet["D6"] = "贸易方式（0110）"
    sheet["D7"] = "一般贸易"
    sheet["H6"] = "征免性质（101）"
    sheet["H7"] = "一般征税"


def test_coded_label_extracts_kv_and_maps_anchor() -> None:
    document = parse_excel(
        _workbook({"报关单": _duoke_label_sheet}),
        file_id="dk",
        filename="duoke.xlsx",
    )
    draft = document.sheets[0]
    assert draft.role == "draft"
    pairs = {item.key: item.value for item in draft.key_values}
    # KV 键保留格子原文，剥码只发生在匹配侧
    assert pairs.get("贸易方式（0110）") == "一般贸易"
    assert pairs.get("征免性质（101）") == "一般征税"
    by_name = _by_name(map_sheet_head(draft, document))
    assert by_name["supvModeCdde"].value == "一般贸易"
    assert by_name["cutMode"].value == "一般征税"


# ---- 繁体 / 多空格键 ----


def _donnelley_invoice_sheet(sheet) -> None:
    sheet["A1"] = "出     口    发     票"
    sheet["H4"] = "发票号:"
    sheet["K4"] = "R26JU551-Y"
    sheet["G7"] = "毛    重:"
    sheet["I7"] = "19137"
    sheet["G8"] = "淨    重:"
    sheet["I8"] = "17432"
    sheet["A9"] = "贸易方式:"
    sheet["C9"] = "一般贸易"
    sheet["G9"] = "件    数:"
    sheet["I9"] = "1278"
    sheet["G10"] = "包裝种类:"
    sheet["I10"] = "纸制或纤维板制盒/箱"


def test_traditional_spaced_keys_map_head_fields() -> None:
    document = parse_excel(
        _workbook({"出口发票": _donnelley_invoice_sheet}),
        file_id="dnn",
        filename="donnelley.xlsx",
    )
    invoice = document.sheets[0]
    by_name = _by_name(map_sheet_head(invoice, document))
    assert by_name["grossWt"].value == "19137"
    assert by_name["netWt"].value == "17432"
    assert by_name["supvModeCdde"].value == "一般贸易"
    assert by_name["packNo"].value == "1278"
    assert by_name["wrapType"].value == "纸制或纤维板制盒/箱"


# ---- 纯代码值路由 ----


def _gsc_party_sheet(sheet) -> None:
    sheet["A1"] = "出境备案清单"
    sheet["A3"] = "经营单位"
    sheet["C3"] = "440356K004"
    sheet["A5"] = "生产销售单位"
    sheet["C5"] = "（440356K004）"


def _credit_code_sheet(sheet) -> None:
    sheet["A1"] = "中华人民共和国海关出口货物报关单"
    sheet["A3"] = "境内发货人"
    sheet["C3"] = "91330206MA2818T42Q"


def test_pure_code_values_route_to_code_fields() -> None:
    document = parse_excel(
        _workbook({"出境备案清单": _gsc_party_sheet}),
        file_id="gsc",
        filename="gsc.xlsx",
    )
    sheet = document.sheets[0]
    by_name = _by_name(map_sheet_head(sheet, document))
    # 10 位海关码 → tradeCode，name 留空待补
    assert by_name["tradeName"].value == ""
    assert by_name["tradeName"].status.value == "needs_review"
    assert "pure_code_value" in by_name["tradeName"].validation_errors
    assert by_name["tradeCode"].value == "440356K004"
    # （码）括号壳剥壳按码处理
    assert by_name["ownerCode"].value == "440356K004"
    assert by_name["ownerName"].status.value == "needs_review"


def test_credit_code_routes_to_scc() -> None:
    document = parse_excel(
        _workbook({"一般贸易出口": _credit_code_sheet}),
        file_id="scc",
        filename="scc.xlsx",
    )
    by_name = _by_name(map_sheet_head(document.sheets[0], document))
    assert by_name["tradeScc"].value == "91330206MA2818T42Q"
    assert by_name["tradeName"].value == ""
    assert by_name["tradeName"].status.value == "needs_review"


def test_name_plus_trailing_code_still_splits() -> None:
    def sheet_fill(sheet) -> None:
        sheet["A1"] = "中华人民共和国海关出口货物报关单"
        sheet["A3"] = "境内发货人"
        sheet["A4"] = "惠州市恒德信精密科技有限公司441394164D"

    document = parse_excel(
        _workbook({"一般贸易出口": sheet_fill}),
        file_id="hx",
        filename="hx.xlsx",
    )
    by_name = _by_name(map_sheet_head(document.sheets[0], document))
    assert by_name["tradeName"].value == "惠州市恒德信精密科技有限公司"
    assert by_name["tradeCode"].value == "441394164D"


# ---- 商品列映射：常量列降级 + 锚点顺序 ----


def _tonda2_sheet(sheet) -> None:
    sheet["A1"] = "申报海关"
    sheet["B1"] = "序号"
    sheet["C1"] = "申报数量"
    sheet["D1"] = "申报计量单位"
    sheet["E1"] = "申报单价"
    sheet["F1"] = "申报总价"
    sheet["G1"] = "净重(千克)"
    sheet["H1"] = "毛重(千克)"
    sheet["I1"] = "总净重(千克)"
    sheet["J1"] = "总毛重(千克)"
    rows = [
        ("1", "3000", "个", "0.0318", "95.49", "2.7", "5.5", "1793.6065", "2045.266"),
        ("2", "18000", "个", "0.0166", "298.02", "10.62", "16.9", "", ""),
        ("3", "4000", "个", "0.0143", "57.2", "1.92", "1.93", "", ""),
    ]
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            sheet.cell(r, c + 1, value)  # B 起填，申报海关列留空


def test_constant_total_column_loses_to_row_column() -> None:
    document = parse_excel(
        _workbook({"Sheet1": _tonda2_sheet}),
        file_id="td2",
        filename="tonda2.xlsx",
    )
    items = map_sheet_goods(document.sheets[0], document)
    assert [item.value_of("gno") for item in items] == ["1", "2", "3"]
    assert [item.value_of("customNetWt") for item in items] == ["2.7", "10.62", "1.92"]
    assert [item.value_of("customGrossWet") for item in items] == ["5.5", "16.9", "1.93"]
    assert items[0].value_of("gqty") == "3000"
    assert items[0].value_of("declPrice") == "0.0318"
    assert items[0].value_of("declTotal") == "95.49"


def _varying_total_sheet(sheet) -> None:
    """国光形：总净重每行不同，是正确列——锚点顺序（总净重更专）应当赢。"""
    sheet["A1"] = "物料名称"
    sheet["B1"] = "净重/个"
    sheet["C1"] = "总净重"
    rows = [
        ("贴纸", "1.3e-05", "0.0013"),
        ("音盆组", "0.001", "22.4"),
        ("防尘帽", "4e-05", "0.84"),
    ]
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            sheet.cell(r, c, value)


def test_varying_total_column_wins_by_anchor_order() -> None:
    document = parse_excel(
        _workbook({"总箱单": _varying_total_sheet}),
        file_id="gg",
        filename="gg.xlsx",
    )
    items = map_sheet_goods(document.sheets[0], document)
    assert [item.value_of("customNetWt") for item in items] == ["0.0013", "22.4", "0.84"]


def _element_vs_spec_sheet(sheet) -> None:
    sheet["A1"] = "装箱单 PACKING LIST"
    sheet["A3"] = "商品名称"
    sheet["B3"] = "商品规格"
    sheet["C3"] = "申报要素"
    sheet["A4"] = "电容"
    sheet["B4"] = "瓷介质"
    sheet["C4"] = "0|0|单层片式|陶瓷|无品牌|无型号"


def test_anchor_order_prefers_specific_declare_element() -> None:
    document = parse_excel(
        _workbook({"装箱单": _element_vs_spec_sheet}),
        file_id="el",
        filename="el.xlsx",
    )
    items = map_sheet_goods(document.sheets[0], document)
    assert items[0].value_of("gmodel") == "0|0|单层片式|陶瓷|无品牌|无型号"


# ---- 码表反查兜底 ----


def test_known_code_reverse_lookup() -> None:
    codes = load_code_tables()
    assert codes.known_code("海关口岸代码", "5304")
    assert not codes.known_code("海关口岸代码", "9999")
    assert not codes.lookup("海关口岸代码", "5304")  # 名称侧查不到


def test_port_code_value_stays_accepted() -> None:
    def sheet_fill(sheet) -> None:
        sheet["A1"] = "出境备案清单"
        sheet["A2"] = "出口口岸"
        sheet["C2"] = "5304"
        sheet["A3"] = "成交方式"
        sheet["C3"] = "FOB"

    document = parse_excel(
        _workbook({"出境备案清单": sheet_fill}),
        file_id="gsc",
        filename="gsc.xlsx",
    )
    declaration = assemble_declaration(document)
    field = declaration.head["iePort"]
    assert field.value == "5304"
    assert field.status.value == "accepted"
    assert not field.validation_errors


# ---- 别名目录 ----


def test_issue66_aliases_present() -> None:
    schema = load_schema()
    anchors = {spec.name: spec.anchors for spec in (*schema.head, *schema.goods)}
    assert "经营单位" in anchors["tradeName"]
    assert "包装类型" in anchors["wrapType"]
    assert "包裝种类" in anchors["wrapType"]
    assert "申报海关" in anchors["customMaster"]
    assert "账册号" in anchors["manualNo"]
    assert "合約號碼" in anchors["contrNo"]
    assert "貿易方式" in anchors["supvModeCdde"]
    assert "淨重" in anchors["netWt"]
    assert "序号" in anchors["gno"]
    assert "申报数量" in anchors["gqty"]
    assert "申报计量单位" in anchors["gunit"]
    assert "申报单价" in anchors["declPrice"]
    assert "申报总价" in anchors["declTotal"]
    assert "金额" in anchors["declTotal"]
    assert "目的国" in anchors["destinationCountry"]


def test_vocab_gained_synced_aliases() -> None:
    vocab = load_layout_vocab()
    groups = {group.id: group for group in (*vocab.box, *vocab.kv)}
    box_texts = {alias.text for group in vocab.box for alias in group.aliases}
    assert "申报海关" in box_texts
    assert "账册号" in box_texts
    assert "貿易方式" in box_texts
    assert "包裝种类" in box_texts
    assert "淨    重" in box_texts
    assert any(alias.text == "经营单位" for alias in groups["trade_party"].aliases)


# ---- 真机样本（本地有才跑） ----


@pytest.mark.skipif(not REAL_GSC.exists(), reason="本地 GSC 样本不在 CI")
def test_gsc_sample_pure_code_routing() -> None:
    document = parse_excel(
        REAL_GSC.read_bytes(),
        file_id="gsc-real",
        filename=REAL_GSC.name,
    )
    declaration = assemble_declaration(document)
    assert declaration.head["tradeCode"].value == "440356K004"
    assert declaration.head["tradeName"].value == ""
    assert declaration.head["tradeName"].status.value == "needs_review"
    assert declaration.head["iePort"].value == "5304"
    assert declaration.head["iePort"].status.value == "accepted"
    assert declaration.head["packNo"].value == "230.000000"
    assert declaration.head["grossWt"].value == "1560"
    assert declaration.head["netWt"].value == "785.02760"
    assert declaration.head["transMode"].value == "FOB"


@pytest.mark.skipif(not REAL_GSRUA.exists(), reason="本地 GSRUA 样本不在 CI")
def test_gsrua_sample_wrap_type_alias() -> None:
    document = parse_excel(
        REAL_GSRUA.read_bytes(),
        file_id="gsrua-real",
        filename=REAL_GSRUA.name,
    )
    declaration = assemble_declaration(document)
    assert declaration.head["wrapType"].value == "胶合板箱,纸箱"
    assert declaration.head["packNo"].value == "10"


@pytest.mark.skipif(
    sys.modules.get("xlrd") is None or not REAL_DUOKE.exists(),
    reason="需要 xlrd 且本地多科样本不在 CI",
)
def test_duoke_sample_coded_labels_and_weights() -> None:
    document = parse_excel(
        REAL_DUOKE.read_bytes(),
        file_id="dk-real",
        filename=REAL_DUOKE.name,
    )
    declaration = assemble_declaration(document)
    assert declaration.head["supvModeCdde"].value == "一般贸易"
    assert declaration.head["cutMode"].value == "一般征税"
    assert declaration.head["grossWt"].value == "625.45"
    assert declaration.head["netWt"].value == "573.1"
    assert declaration.head["packNo"].value == "42"
    assert declaration.head["wrapType"].value == "纸箱/其他包装"
    assert declaration.head["transMode"].value == "FOB"


@pytest.mark.skipif(
    sys.modules.get("xlrd") is None or not REAL_DONNELLEY.exists(),
    reason="需要 xlrd 且本地当纳利样本不在 CI",
)
def test_donnelley_sample_traditional_keys() -> None:
    document = parse_excel(
        REAL_DONNELLEY.read_bytes(),
        file_id="dnn-real",
        filename=REAL_DONNELLEY.name,
    )
    invoice = next(sheet for sheet in document.sheets if sheet.name == "出口发票")
    by_name = _by_name(map_sheet_head(invoice, document))
    assert by_name["grossWt"].value == "19137"
    assert by_name["netWt"].value == "17432"


@pytest.mark.skipif(not REAL_TONDA2.exists(), reason="本地通达2 样本不在 CI")
def test_tonda2_sample_row_level_weights() -> None:
    document = parse_excel(
        REAL_TONDA2.read_bytes(),
        file_id="td2-real",
        filename=REAL_TONDA2.name,
    )
    items = map_sheet_goods(document.sheets[0], document)
    assert len(items) == 50
    assert items[0].value_of("customNetWt") == "2.7"
    assert items[1].value_of("customNetWt") == "10.62"
    assert items[2].value_of("customNetWt") == "1.92"
    assert items[0].value_of("gno") == "1"
    assert items[0].value_of("customGrossWet") == "5.5"
