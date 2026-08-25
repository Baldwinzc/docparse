"""平表表列 → 表头字段（#67）。

覆盖：恒定列出字段、每行变化列不出、合计只填首行、单行数据 review、
表尾冒号注记行走 same_cell、框表角色不走列路径、真机通达2。
"""

from __future__ import annotations

import io
from pathlib import Path

import pytest

pytest.importorskip("openpyxl")

from openpyxl import Workbook
from openpyxl.styles import Border, Side

from docparse.adapters.parsers.excel import parse_excel
from docparse.extraction.head_map import map_sheet_head
from docparse.schema.loader import load_schema, load_sheet_roles

_SUPPLY = Path("/Users/baldwin/Desktop/taizhou/补充测试")
REAL_TONGDA = _SUPPLY / "通达2.xlsx"

_HEADERS = [
    "申报海关",
    "账册号",
    "申报计量单位",
    "第一法定数量",
    "总件数",
    "总净重(千克)",
    "总毛重(千克)",
    "序号",
    "商品编码",
    "商品名称",
    "申报数量",
    "件数",
    "净重(千克)",
    "毛重(千克)",
    "申报日期",
    "出口口岸",
    "监管方式",
    "征免性质",
    "包装类型",
    "成交方式",
    "贸易国(地区)",
    "指运港",
    "离境口岸",
]


def _thin() -> Border:
    line = Side(style="thin")
    return Border(left=line, right=line, top=line, bottom=line)


def _workbook(builders: dict) -> bytes:
    book = Workbook()
    first = True
    for title, fill in builders.items():
        sheet = book.active if first else book.create_sheet(title)
        if first:
            sheet.title = title
            first = False
        fill(sheet)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _flat_rows(total_first_only: bool = False) -> list[list]:
    rows = [
        ["东兴海关", "H1234", "个", "10", "272", "1793.6", "2045.2",
         "1", "8532221000", "电容", "3000", "4", "2.7", "4.6",
         "2026/08/13", "东兴海关", "一般贸易", "一般征税", "纸制或纤维板制盒/箱",
         "FOB", "越南", "越南", "东兴"],
        ["东兴海关", "H1234", "个", "5", "272", "1793.6", "2045.2",
         "2", "8532221000", "电容", "1500", "2", "1.3", "2.2",
         "2026/08/13", "东兴海关", "一般贸易", "一般征税", "纸制或纤维板制盒/箱",
         "FOB", "越南", "越南", "东兴"],
        ["东兴海关", "H1234", "个", "8", "272", "1793.6", "2045.2",
         "3", "8544422900", "喇叭线", "3900", "6", "3.4", "5.1",
         "2026/08/13", "东兴海关", "一般贸易", "一般征税", "纸制或纤维板制盒/箱",
         "FOB", "越南", "越南", "东兴"],
    ]
    if total_first_only:
        # 合计只填首行（通达2 真实形状）
        for row in rows[1:]:
            for col in (4, 5, 6):  # 总件数 / 总净重 / 总毛重
                row[col] = ""
    return rows


def _flat_sheet(rows: list[list], notes: bool = True) -> None:
    def fill(sheet) -> None:
        for index, header in enumerate(_HEADERS, start=1):
            sheet.cell(1, index, header)
        for r, row in enumerate(rows, start=2):
            for c, value in enumerate(row, start=1):
                if value != "":
                    sheet.cell(r, c, value)
        last = len(rows) + 1
        if notes:
            sheet.cell(last + 1, 1, "境内发货人:惠州市通达国际供应链管理有限公司")
            sheet.cell(last + 2, 1, "境外收货人:TONLY ELECTRONICS TECHNOLOGY VIET NAM")
            last += 2
        for row in sheet.iter_rows(min_row=1, max_row=last, min_col=1, max_col=len(_HEADERS)):
            for cell in row:
                cell.border = _thin()

    return fill


def _draft_with_constant_column(sheet) -> None:
    """框表草单：商品表带恒定「监管方式」列，但 KV 里没有监管方式。

    draft 未声明 head_from_columns，列路径必须不开，supvModeCdde 不得出现。
    """
    sheet["A1"] = "中华人民共和国海关出口货物报关单"
    sheet["A3"] = "境内发货人"
    sheet["A4"] = "惠州市恒德信精密科技有限公司"
    sheet["A5"] = "境外收货人"
    sheet["A6"] = "BLU PRECISION LIMITED"
    goods_headers = ["项号", "商品编号", "商品名称及规格型号", "监管方式"]
    for index, header in enumerate(goods_headers, start=1):
        sheet.cell(17, index, header)
    for row_index, item_no in enumerate((1, 2, 3), start=18):
        sheet.cell(row_index, 1, item_no)
        sheet.cell(row_index, 2, "9111900000")
        sheet.cell(row_index, 3, "电容")
        sheet.cell(row_index, 4, "一般贸易")
    for row in sheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=4):
        for cell in row:
            cell.border = _thin()


def _parse(fill, *, title="Sheet1", file_id="fx", filename="fixture.xlsx"):
    document = parse_excel(_workbook({title: fill}), file_id=file_id, filename=filename)
    return document, document.sheets[0]


def _by_name(fields) -> dict:
    return {item.name: item for item in fields}


def test_sheet_roles_flag_declaration_list_head_from_columns() -> None:
    roles = load_sheet_roles()
    role = roles.role("declaration_list")
    assert role is not None and role.head_from_columns
    draft = roles.role("draft")
    assert draft is not None and not draft.head_from_columns


def test_constant_columns_map_head_fields() -> None:
    document, sheet = _parse(_flat_sheet(_flat_rows()))
    assert sheet.role == "declaration_list"
    by_name = _by_name(map_sheet_head(sheet, document))
    assert by_name["customMaster"].value == "东兴海关"
    assert by_name["iePort"].value == "东兴海关"
    assert by_name["declDate"].value == "2026/08/13"
    assert by_name["supvModeCdde"].value == "一般贸易"
    assert by_name["cutMode"].value == "一般征税"
    assert by_name["wrapType"].value == "纸制或纤维板制盒/箱"
    assert by_name["transMode"].value == "FOB"
    assert by_name["cusTradeNationCode"].value == "越南"
    assert by_name["distinatePort"].value == "越南"
    assert by_name["ciqEntyPortCode"].value == "东兴"
    assert by_name["packNo"].value == "272"
    assert by_name["grossWt"].value == "2045.2"
    assert by_name["netWt"].value == "1793.6"
    for name in ("packNo", "grossWt", "netWt", "declDate", "customMaster"):
        field = by_name[name]
        assert field.status.value == "accepted"
        assert field.evidence[0].cell.startswith("Sheet1!")
        assert field.evidence[0].quote.startswith("Sheet1!")


def test_totals_written_only_on_first_row_still_map() -> None:
    """合计列只填首行（通达2 真实形状）：无冲突值即恒定。"""
    document, sheet = _parse(_flat_sheet(_flat_rows(total_first_only=True)))
    by_name = _by_name(map_sheet_head(sheet, document))
    assert by_name["packNo"].value == "272"
    assert by_name["grossWt"].value == "2045.2"
    assert by_name["netWt"].value == "1793.6"
    assert by_name["packNo"].status.value == "accepted"


def test_variable_columns_do_not_map() -> None:
    """每行变化的件数 / 净重 / 毛重是行级信息，不出表头。"""
    document, sheet = _parse(_flat_sheet(_flat_rows()))
    by_name = _by_name(map_sheet_head(sheet, document))
    assert by_name["packNo"].value == "272"  # 取总件数，不是行级 4/2/6
    assert by_name["netWt"].value == "1793.6"
    assert by_name["grossWt"].value == "2045.2"


def test_single_row_table_marks_needs_review() -> None:
    document, sheet = _parse(_flat_sheet(_flat_rows()[:1]))
    by_name = _by_name(map_sheet_head(sheet, document))
    assert by_name["declDate"].value == "2026/08/13"
    assert by_name["declDate"].status.value == "needs_review"
    assert "single_row_column" in by_name["declDate"].validation_errors


def test_trailing_note_rows_become_same_cell_kv() -> None:
    """表尾冒号注记行不进表体，走 same_cell：tradeName / consignorEname 有值。"""
    document, sheet = _parse(_flat_sheet(_flat_rows()))
    table = sheet.tables[0]
    assert len(table.rows) == 3  # 注记行不在表体
    pairs = {pair.key: pair for pair in sheet.key_values}
    assert pairs["境内发货人"].value == "惠州市通达国际供应链管理有限公司"
    assert pairs["境内发货人"].strategy == "same_cell"
    assert pairs["境外收货人"].value.startswith("TONLY ELECTRONICS")
    by_name = _by_name(map_sheet_head(sheet, document))
    assert by_name["tradeName"].value == "惠州市通达国际供应链管理有限公司"
    assert by_name["consignorEname"].value.startswith("TONLY ELECTRONICS")


def test_draft_role_never_reads_columns() -> None:
    """框表角色不声明 head_from_columns：商品表常量列不出表头字段。"""
    document, sheet = _parse(_draft_with_constant_column, title="一般贸易出口")
    assert sheet.role == "draft"
    by_name = _by_name(map_sheet_head(sheet, document))
    assert "supvModeCdde" not in by_name
    assert by_name["tradeName"].value == "惠州市恒德信精密科技有限公司"


def test_schema_total_anchors_registered() -> None:
    schema = load_schema()
    assert "总件数" in schema.field("packNo").anchors
    assert "总毛重（千克）" in schema.field("grossWt").anchors
    assert "总毛重(千克)" in schema.field("grossWt").anchors
    assert "总净重（千克）" in schema.field("netWt").anchors
    assert "总净重(千克)" in schema.field("netWt").anchors
    assert "申报海关" in schema.field("customMaster").anchors


@pytest.mark.skipif(not REAL_TONGDA.exists(), reason="本地通达2 样本不在 CI")
def test_real_tongda_head_from_columns() -> None:
    document = parse_excel(
        REAL_TONGDA.read_bytes(),
        file_id="td-real",
        filename=REAL_TONGDA.name,
    )
    sheet = document.sheets[0]
    assert sheet.role == "declaration_list"
    by_name = _by_name(map_sheet_head(sheet, document))
    assert by_name["declDate"].value == "2026/08/13"
    assert by_name["iePort"].value == "东兴海关"
    assert by_name["customMaster"].value == "东兴海关"
    assert by_name["supvModeCdde"].value == "一般贸易"
    assert by_name["cutMode"].value == "一般征税"
    assert by_name["wrapType"].value == "纸制或纤维板制盒/箱"
    assert by_name["transMode"].value == "FOB"
    assert by_name["cusTradeNationCode"].value == "越南"
    assert by_name["cusTradeCountry"].value == "越南"
    assert by_name["distinatePort"].value == "越南"
    assert by_name["ciqEntyPortCode"].value == "东兴"
    assert by_name["packNo"].value == "272"
    assert by_name["grossWt"].value == "2045.266"
    assert by_name["netWt"].value == "1793.6065"
    assert by_name["tradeName"].value == "惠州市通达国际供应链管理有限公司"
    assert by_name["consignorEname"].value.startswith("TONLY ELECTRON")
    assert by_name["packNo"].status.value == "accepted"
