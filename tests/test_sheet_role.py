from __future__ import annotations

import io
from pathlib import Path

import pytest

pytest.importorskip("openpyxl")

from openpyxl import Workbook
from openpyxl.styles import Border, Side

from docparse.adapters.parsers.excel import parse_excel
from docparse.schema.loader import load_sheet_roles

_DEMO = Path("/Users/baldwin/Desktop/taizhou/AI识别Demo")
REAL_HENGXIN = _DEMO / "（恒信）一般贸易草单HDX260251BLU.xlsx"
REAL_GUOGUANG = _DEMO / "（国光）箱单发票合同26VN0502-1.xlsx"


def _thin() -> Border:
    line = Side(style="thin")
    return Border(left=line, right=line, top=line, bottom=line)


def _workbook(builders: dict) -> bytes:
    book = Workbook()
    first = True
    for title, fill in builders.items():
        sheet = book.active if first else book.create_sheet(title)
        if first:
            sheet.title = title
            first = False
        fill(sheet)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _draft(sheet) -> None:
    sheet["A1"] = "中华人民共和国海关出口货物报关单"
    sheet["A3"] = "境内发货人"
    sheet["A4"] = "示例公司"
    sheet["E3"] = "出境关别"
    sheet["E4"] = "莲塘口岸"
    sheet["E7"] = "监管方式"
    sheet["E8"] = "一般贸易"
    sheet["I7"] = "征免性质"
    sheet["I8"] = "一般征税"
    sheet["A17"] = "项号"
    sheet["B17"] = "商品编号"
    sheet["C17"] = "商品名称及规格型号"
    sheet["D17"] = "申报要素"
    sheet["A18"] = 1
    sheet["B18"] = "9111900000"
    sheet["C18"] = "壳体"
    sheet["D18"] = "不锈钢"
    for row in sheet.iter_rows(min_row=1, max_row=18, min_col=1, max_col=4):
        for cell in row:
            cell.border = _thin()


def _packing(sheet) -> None:
    sheet["A1"] = "PACKING LIST"
    sheet["I4"] = "Invoice No.﹕"
    sheet["J4"] = "HDX260251"
    sheet["A7"] = "Bill To : BLU PRECISION LIMITED"
    sheet["C11"] = "Description (货物名称)"
    sheet["E11"] = "Ship Q'ty (数量/PCS)"
    sheet["H11"] = "N.W. (Kg) (净重)"
    sheet["I11"] = "G.W .(Kg) (毛重)"
    sheet["C12"] = "壳体"
    sheet["E12"] = 150
    sheet["H12"] = 5.74
    sheet["I12"] = 7.35
    for row in sheet.iter_rows(min_row=1, max_row=12, min_col=1, max_col=10):
        for cell in row:
            cell.border = _thin()


def _invoice(sheet) -> None:
    sheet["A1"] = "INVOICE"
    sheet["F6"] = "发票号INVOICE NO.:"
    sheet["G6"] = "26VN0502-1"
    sheet["A11"] = "运输工具 SHIPPED PER:"
    sheet["B11"] = "BY TRUCK"
    sheet["B16"] = "物料名称"
    sheet["D16"] = "出货数量"
    sheet["F16"] = "单价"
    sheet["G16"] = "汇总价"
    sheet["B17"] = "贴纸"
    sheet["D17"] = 100
    for row in sheet.iter_rows(min_row=1, max_row=17, min_col=1, max_col=7):
        for cell in row:
            cell.border = _thin()


def _contract(sheet) -> None:
    sheet["C1"] = "合     同"
    sheet["C2"] = "SALES CONTRACT"
    sheet["A5"] = "卖 方:"
    sheet["B5"] = "示例卖方"
    sheet["A8"] = "买 方:"
    sheet["B8"] = "示例买方"
    sheet["F6"] = "合同号："
    sheet["G6"] = "HDX2026-251"
    sheet["A14"] = "(1)货物名称"
    sheet["C14"] = "(3)数量"
    sheet["F14"] = "(6)总值"
    sheet["A15"] = "Description"
    sheet["C15"] = "Quantity"
    sheet["F15"] = "Total Amount(HKD)"
    sheet["A16"] = 1
    sheet["C16"] = 150
    sheet["F16"] = 14700
    for row in sheet.iter_rows(min_row=1, max_row=16, min_col=1, max_col=7):
        for cell in row:
            cell.border = _thin()


def _code_lookup(sheet) -> None:
    sheet["C2"] = "BLU PRECISION LIMITED"
    sheet["G2"] = "005"
    sheet["C3"] = "TIME INVENTION LIMITED"
    sheet["G3"] = "006"
    sheet["C4"] = "HANSON PRECISION TECHNOLOGY LIMITED"
    sheet["G4"] = "002"


def _internal_sku(sheet) -> None:
    sheet["B1"] = "编号"
    sheet["C1"] = "货物名称"
    sheet["E1"] = "恒信编号"
    sheet["F1"] = "数量"
    sheet["O1"] = "ERP编号"
    sheet["B2"] = "D001"
    sheet["C2"] = "表殻殻身"
    sheet["E2"] = "HNC3591GA"
    sheet["F2"] = 150
    for row in sheet.iter_rows(min_row=1, max_row=2, min_col=2, max_col=15):
        for cell in row:
            cell.border = _thin()


def _history_ledger(sheet) -> None:
    sheet["A1"] = "序号"
    sheet["B1"] = "报关日期"
    sheet["C1"] = "报关单号"
    sheet["H1"] = "商品编码"
    sheet["K1"] = "SAP编号"
    sheet["M1"] = "商品名称"
    sheet["A2"] = 1
    sheet["C2"] = "E123"
    sheet["H2"] = "9111900000"
    for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=13):
        for cell in row:
            cell.border = _thin()


def _guoguang_packing(sheet) -> None:
    sheet["E5"] = "装箱单 PACKING LIST"
    sheet["A6"] = "日期DATE:"
    sheet["B6"] = "2026-03-15"
    sheet["A7"] = "发票/INVOICE NO.:"
    sheet["B7"] = "26VN0502-1"
    sheet["G7"] = "合同号CONTRACT NO.:"
    sheet["H7"] = "26VN0502"
    sheet["B12"] = "物料名称"
    sheet["D12"] = "出货数量"
    sheet["F12"] = "总箱数"
    sheet["H12"] = "总净重 NW"
    sheet["J12"] = "总毛重 GW"
    sheet["B13"] = "贴纸"
    sheet["D13"] = 100
    for row in sheet.iter_rows(min_row=5, max_row=13, min_col=1, max_col=10):
        for cell in row:
            cell.border = _thin()


def _notes(sheet) -> None:
    sheet["A1"] = "临时备注"
    sheet["A2"] = "今天对一下件数"
    sheet["A3"] = "不要当箱单"


def _declaration_list(sheet) -> None:
    headers = [
        "申报日期",
        "申报海关",
        "账册号",
        "备案序号",
        "申报计量单位",
        "第一法定数量",
        "第二法定单位",
        "申报单价",
        "申报总价",
        "总件数",
        "总净重(千克)",
        "总毛重(千克)",
        "申报要素",
        "监管方式",
        "征免性质",
        "成交方式",
        "包装类型",
        "贸易国(地区)",
        "运抵国(地区)",
        "指运港",
        "离境口岸",
        "境内货源地",
    ]
    for index, header in enumerate(headers, start=1):
        sheet.cell(1, index, header)
    sheet.cell(2, 1, "2026/08/13")
    sheet.cell(2, 2, "东兴海关")
    sheet.cell(2, 3, "44133607NZ-001")
    for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = _thin()


def _gsc_draft(sheet) -> None:
    sheet["A1"] = "出境备案清单"
    sheet["A2"] = "出口口岸"
    sheet["C2"] = "5304"
    sheet["F2"] = "备案号"
    sheet["G2"] = "T5339W000208"
    sheet["A3"] = "运输方式"
    sheet["B3"] = "水路运输"
    sheet["A4"] = "贸易方式"
    sheet["B4"] = "区内物流货物"
    sheet["A5"] = "境外收货人"
    sheet["B5"] = "GREAT SUN WAY ENTERPRISE LTD."
    sheet["A6"] = "运抵国"
    sheet["B6"] = "中国台湾"
    sheet["C6"] = "指运港"
    sheet["D6"] = "中国台湾"
    sheet["E6"] = "离境口岸"
    sheet["F6"] = "赤湾"
    sheet["A7"] = "成交方式"
    sheet["B7"] = "FOB"
    sheet["C7"] = "包装种类"
    sheet["D7"] = "纸箱"
    sheet["A8"] = "项号"
    sheet["B8"] = "海关编码"
    sheet["A9"] = 1
    sheet["B9"] = "8528521200"
    for row in sheet.iter_rows(min_row=1, max_row=9, min_col=1, max_col=6):
        for cell in row:
            cell.border = _thin()


def _dangnali_packing(sheet) -> None:
    sheet["H1"] = "装 箱 单"
    sheet["A6"] = "项目号"
    sheet["F6"] = "数量 (本)"
    sheet["G6"] = "净重 (KG)"
    sheet["H6"] = "毛重 (KG)"
    sheet["J6"] = "纸箱数"
    sheet["A7"] = "P001"
    sheet["F7"] = 100
    for row in sheet.iter_rows(min_row=6, max_row=7, min_col=1, max_col=10):
        for cell in row:
            cell.border = _thin()


def _dangnali_invoice(sheet) -> None:
    sheet["A2"] = "出     口    发     票"
    sheet["H4"] = "发票号:"
    sheet["K4"] = "R26JU551-Y"
    sheet["A10"] = "序号"
    sheet["B10"] = "商品编号"
    sheet["C10"] = "商品名称"
    sheet["A11"] = 1
    sheet["B11"] = "4901990000"
    for row in sheet.iter_rows(min_row=10, max_row=11, min_col=1, max_col=3):
        for cell in row:
            cell.border = _thin()


def _dangnali_contract(sheet) -> None:
    sheet["A1"] = "销 售 合 同"
    sheet["A4"] = "卖方(SELLER)："
    sheet["B4"] = "示例卖方"
    sheet["F4"] = "买方(BUYER)："
    sheet["G4"] = "示例买方"
    sheet["A8"] = "卖方"
    sheet["B8"] = "示例卖方"
    sheet["F8"] = "买方"
    sheet["G8"] = "示例买方"
    sheet["A10"] = "商品名称"
    sheet["B10"] = "数量"
    sheet["C10"] = "单价"
    sheet["D10"] = "总价"
    sheet["A11"] = "书刊"
    sheet["B11"] = 10
    for row in sheet.iter_rows(min_row=10, max_row=11, min_col=1, max_col=4):
        for cell in row:
            cell.border = _thin()


def _mxy_contract(sheet) -> None:
    sheet["A1"] = "合     同"
    sheet["A3"] = "CONTRACT"
    sheet["A4"] = "賣  方"
    sheet["B4"] = "示例卖方"
    sheet["F4"] = "合約號碼"
    sheet["G4"] = "MXY2026-0616"
    sheet["A8"] = "買  方"
    sheet["B8"] = "示例买方"
    sheet["A12"] = "Name of commodity and Specification"
    sheet["B12"] = "Quantity"
    sheet["C12"] = "Unit Price"
    sheet["D12"] = "Amount"
    sheet["A13"] = "UPS"
    sheet["B13"] = 8
    for row in sheet.iter_rows(min_row=12, max_row=13, min_col=1, max_col=4):
        for cell in row:
            cell.border = _thin()


def _duoke_hidden(sheet) -> None:
    sheet["A1"] = "型号"
    sheet["B1"] = "报关型号"
    sheet["C1"] = "报关单价"
    sheet["D1"] = "申报要素"
    sheet["A2"] = "DCM8"
    sheet["B2"] = "DCM8"
    sheet["C2"] = 158
    sheet["D2"] = "1|0|电脑显示拓展用"
    for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=4):
        for cell in row:
            cell.border = _thin()


def _elements_sheet(sheet) -> None:
    sheet["A1"] = "规格型号"
    sheet["C1"] = "申报要素"
    sheet["A2"] = "DCM8"
    sheet["C2"] = "1|0|电脑显示拓展用"


def _roles(document) -> dict[str, str]:
    return {sheet.name: sheet.role for sheet in document.sheets}


def test_sheet_roles_catalog_loads() -> None:
    catalog = load_sheet_roles()
    assert {role.id for role in catalog.roles} == {
        "draft",
        "declaration_list",
        "packing",
        "invoice",
        "contract",
        "auxiliary",
    }
    assert catalog.role("draft").consume == "primary"
    assert catalog.role("declaration_list").consume == "primary"
    assert catalog.role("packing").consume == "supplement"
    assert catalog.role("auxiliary").consume == "exclude"
    assert catalog.unknown_consume == "exclude"
    assert all(
        signal.source
        for role in catalog.roles
        for group in (
            role.signals.titles,
            role.signals.keys,
            role.signals.headers,
            role.signals.filename,
        )
        for signal in group
    )


def test_fixture_covers_hengxin_seven_roles() -> None:
    document = parse_excel(
        _workbook(
            {
                "一般贸易出口": _draft,
                "裝箱單": _packing,
                "发票": _invoice,
                "合同": _contract,
                "Sheet1": _code_lookup,
                "1": _internal_sku,
                "Sheet3": _history_ledger,
            }
        ),
        file_id="hx",
        filename="hengxin-fixture.xlsx",
    )
    assert _roles(document) == {
        "一般贸易出口": "draft",
        "裝箱單": "packing",
        "发票": "invoice",
        "合同": "contract",
        "Sheet1": "auxiliary",
        "1": "auxiliary",
        "Sheet3": "auxiliary",
    }
    consume = {sheet.name: sheet.consume for sheet in document.sheets}
    assert consume["一般贸易出口"] == "primary"
    assert consume["裝箱單"] == "supplement"
    assert consume["Sheet1"] == "exclude"
    leftover = next(sheet for sheet in document.sheets if sheet.name == "Sheet1")
    assert leftover.cells
    assert leftover.role == "auxiliary"


def test_fixture_covers_guoguang_packing_and_invoice() -> None:
    document = parse_excel(
        _workbook({"总箱单": _guoguang_packing, "发票 ": _invoice}),
        file_id="gg",
        filename="guoguang-fixture.xlsx",
    )
    assert _roles(document) == {"总箱单": "packing", "发票 ": "invoice"}
    packing = next(sheet for sheet in document.sheets if sheet.name == "总箱单")
    assert packing.consume == "supplement"


def test_nameless_sheet_still_classifies_from_content() -> None:
    document = parse_excel(
        _workbook({"Sheet1": _draft, "工作表2": _packing}),
        file_id="anon",
        filename="untitled.xlsx",
    )
    assert _roles(document) == {"Sheet1": "draft", "工作表2": "packing"}


def test_filename_alone_does_not_route() -> None:
    document = parse_excel(
        _workbook({"Sheet2": _notes}),
        file_id="fn",
        filename="一般贸易草单箱单发票合同.xlsx",
    )
    notes = document.sheets[0]
    assert notes.role == "unknown"
    assert notes.consume == "exclude"
    assert any(cell.value == "临时备注" for cell in notes.cells)


def test_unknown_keeps_cells() -> None:
    document = parse_excel(
        _workbook({"备注": _notes}),
        file_id="unk",
        filename="notes.xlsx",
    )
    sheet = document.sheets[0]
    assert sheet.role == "unknown"
    assert sheet.consume == "exclude"
    assert any(cell.value == "临时备注" for cell in sheet.cells)


@pytest.mark.skipif(not REAL_HENGXIN.exists(), reason="本地恒信样本不在 CI")
def test_hengxin_sample_roles() -> None:
    document = parse_excel(
        REAL_HENGXIN.read_bytes(),
        file_id="hx-real",
        filename=REAL_HENGXIN.name,
    )
    roles = _roles(document)
    assert roles["一般贸易出口"] == "draft"
    assert any(name in {"裝箱單", "装箱单"} and role == "packing" for name, role in roles.items())
    assert any("发票" in name and role == "invoice" for name, role in roles.items())
    assert roles["合同"] == "contract"
    auxiliaries = [name for name, role in roles.items() if role == "auxiliary"]
    assert len(auxiliaries) == 3


@pytest.mark.skipif(not REAL_GUOGUANG.exists(), reason="本地国光样本不在 CI")
def test_guoguang_sample_roles() -> None:
    document = parse_excel(
        REAL_GUOGUANG.read_bytes(),
        file_id="gg-real",
        filename=REAL_GUOGUANG.name,
    )
    roles = _roles(document)
    assert roles["总箱单"] == "packing"
    invoice = next(name for name in roles if "发票" in name)
    assert roles[invoice] == "invoice"


def test_flat_customs_table_is_declaration_list() -> None:
    document = parse_excel(
        _workbook({"Sheet1": _declaration_list}),
        file_id="td",
        filename="tongda-fixture.xlsx",
    )
    sheet = document.sheets[0]
    assert sheet.role == "declaration_list"
    assert sheet.consume == "primary"
    assert sheet.role_confidence >= 0.95


def test_boxed_filing_list_is_draft() -> None:
    document = parse_excel(
        _workbook({"Sheet1": _gsc_draft}),
        file_id="gsc",
        filename="gsc-fixture.xlsx",
    )
    sheet = document.sheets[0]
    assert sheet.role == "draft"
    assert sheet.consume == "primary"


def test_dangnali_style_titles_route_commercial_sheets() -> None:
    document = parse_excel(
        _workbook(
            {
                "装 箱 单": _dangnali_packing,
                "出口发票": _dangnali_invoice,
                "销售合同": _dangnali_contract,
            }
        ),
        file_id="dn",
        filename="dangnali-fixture.xls",
    )
    assert _roles(document) == {
        "装 箱 单": "packing",
        "出口发票": "invoice",
        "销售合同": "contract",
    }


def test_traditional_contract_keys_route_contract() -> None:
    document = parse_excel(
        _workbook({"合同": _mxy_contract}),
        file_id="mxy",
        filename="mxy-fixture.xlsx",
    )
    sheet = document.sheets[0]
    assert sheet.role == "contract"
    assert sheet.consume == "supplement"


def test_elements_and_hidden_stay_non_primary() -> None:
    document = parse_excel(
        _workbook({"申报要素": _elements_sheet, "隐藏": _duoke_hidden}),
        file_id="dk",
        filename="duoke-fixture.xls",
    )
    roles = _roles(document)
    assert roles["申报要素"] in {"unknown", "auxiliary"}
    assert roles["隐藏"] in {"unknown", "auxiliary"}
    consume = {sheet.name: sheet.consume for sheet in document.sheets}
    assert consume["申报要素"] == "exclude"
    assert consume["隐藏"] == "exclude"


def test_hengxin_draft_still_beats_declaration_list() -> None:
    document = parse_excel(
        _workbook({"一般贸易出口": _draft, "Sheet1": _declaration_list}),
        file_id="mix",
        filename="hengxin-fixture.xlsx",
    )
    assert _roles(document) == {
        "一般贸易出口": "draft",
        "Sheet1": "declaration_list",
    }


def test_baoguan_substring_does_not_promote_lookup() -> None:
    document = parse_excel(
        _workbook({"隐藏": _duoke_hidden, "Sheet3": _history_ledger}),
        file_id="aux",
        filename="duoke-fixture.xls",
    )
    assert _roles(document) == {"隐藏": "unknown", "Sheet3": "auxiliary"}
