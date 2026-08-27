from __future__ import annotations

import io
from pathlib import Path

import pytest

pytest.importorskip("openpyxl")

from openpyxl import Workbook
from openpyxl.styles import Border, Side

from docparse.adapters.parsers.excel import parse_excel
from docparse.extraction.goods_map import map_document_goods, map_sheet_goods
from docparse.schema.loader import load_schema

_DEMO = Path("/Users/baldwin/Desktop/taizhou/AI识别Demo")
_SUPPLY = Path("/Users/baldwin/Desktop/taizhou/补充测试")
REAL_HENGXIN = _DEMO / "（恒信）一般贸易草单HDX260251BLU.xlsx"
REAL_GUOGUANG = _DEMO / "（国光）箱单发票合同26VN0502-1.xlsx"
REAL_GSRUA = _SUPPLY / "GSRUA26601CLLG01(1).xlsx"
REAL_DONNELLEY = _SUPPLY / "202606 R26JU551-Y报关一般.xls"
REAL_MXY = _SUPPLY / "MXY2026-0616 东莞-越南物料 一般出口报关资料 UPS 更新.xlsx"
REAL_DUOKE = _SUPPLY / "6-17 多科报关资料香港出货 DKTX-2606057 多科通讯乐乐高(1).xls"
REAL_BLU_IN = _SUPPLY / "进料加工出口报关单-BLU-HDX260271(1).xlsx"


def _thin() -> Border:
    line = Side(style="thin")
    return Border(left=line, right=line, top=line, bottom=line)


def _workbook(builders: dict) -> bytes:
    book = Workbook()
    first = True
    for title, fill in builders.items():
        sheet = book.active if first else book.create_sheet(title)
        if first:
            sheet.title = title
            first = False
        fill(sheet)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _draft(sheet) -> None:
    sheet["A1"] = "中华人民共和国海关出口货物报关单"
    sheet["A3"] = "境内发货人"
    sheet["A4"] = "示例公司"
    sheet["E3"] = "出境关别"
    sheet["E4"] = "莲塘口岸"
    sheet["E7"] = "监管方式"
    sheet["E8"] = "一般贸易"
    sheet["I7"] = "征免性质"
    sheet["I8"] = "一般征税"
    headers = [
        ("A17", "项号"),
        ("B17", "商品编号"),
        ("C17", "商品名称及规格型号"),
        ("D17", "申报要素"),
        ("E17", "数量PCS"),
        ("F17", "计价单位"),
        ("G17", "重量KG"),
        ("H17", "单价"),
        ("I17", "总价"),
        ("J17", "币制"),
        ("K17", "原产国（地区）"),
        ("L17", "最终目的国（地区）"),
        ("M17", "境内货源地"),
    ]
    for address, text in headers:
        sheet[address] = text
    sheet["A18"] = 1
    sheet["B18"] = "9111900000"
    sheet["C18"] = "表壳配件/壳体"
    sheet["D18"] = "不锈钢/无中文品牌、无外文品牌/无型号"
    sheet["E18"] = 150
    sheet["F18"] = "只"
    sheet["G18"] = 5.74
    sheet["H18"] = 98
    sheet["I18"] = 14700
    sheet["J18"] = "港币"
    sheet["K18"] = "中国"
    sheet["L18"] = "中国香港"
    sheet["M18"] = "惠州其他"
    for row in sheet.iter_rows(min_row=1, max_row=18, min_col=1, max_col=13):
        for cell in row:
            cell.border = _thin()


def _packing(sheet) -> None:
    sheet["A1"] = "PACKING LIST"
    sheet["I4"] = "Invoice No.﹕"
    sheet["J4"] = "HDX260251"
    sheet["A7"] = "Bill To : BLU PRECISION LIMITED"
    sheet["C11"] = "Description (货物名称)"
    sheet["E11"] = "Ship Q'ty (数量/PCS)"
    sheet["G11"] = "Packing (箱)"
    sheet["H11"] = "N.W. (Kg) (净重)"
    sheet["I11"] = "G.W .(Kg) (毛重)"
    sheet["C12"] = "表壳配件/壳体"
    sheet["E12"] = 150
    sheet["G12"] = 40
    sheet["H12"] = 5.74
    sheet["I12"] = 7.35
    for row in sheet.iter_rows(min_row=1, max_row=12, min_col=1, max_col=10):
        for cell in row:
            cell.border = _thin()


def _invoice(sheet) -> None:
    sheet["A1"] = "INVOICE"
    sheet["F6"] = "发票号INVOICE NO.:"
    sheet["G6"] = "26VN0502-1"
    sheet["A11"] = "运输工具 SHIPPED PER:"
    sheet["B11"] = "BY TRUCK"
    sheet["B16"] = "物料名称"
    sheet["D16"] = "出货数量"
    sheet["E16"] = "币别"
    sheet["F16"] = "单价"
    sheet["G16"] = "汇总价"
    sheet["B17"] = "贴纸"
    sheet["D17"] = 100
    sheet["E17"] = "USD"
    sheet["F17"] = 0.0181
    sheet["G17"] = 1.81
    for row in sheet.iter_rows(min_row=1, max_row=17, min_col=1, max_col=7):
        for cell in row:
            cell.border = _thin()


def _guoguang_packing(sheet) -> None:
    sheet["E5"] = "装箱单 PACKING LIST"
    sheet["A6"] = "日期DATE:"
    sheet["B6"] = "2026-03-15"
    sheet["A7"] = "发票/INVOICE NO.:"
    sheet["B7"] = "26VN0502-1"
    sheet["G7"] = "合同号CONTRACT NO.:"
    sheet["H7"] = "26VN0502"
    sheet["A8"] = "卖方SELLER"
    sheet["A9"] = "GUOGUANG ELECTRIC COMPANY LIMITED."
    sheet["B12"] = "物料名称"
    sheet["C12"] = "单位"
    sheet["D12"] = "出货数量"
    sheet["H12"] = "总净重 NW"
    sheet["J12"] = "总毛重 GW"
    sheet["Q12"] = "HS CODE海关编码"
    sheet["R12"] = "申报要素"
    sheet["S12"] = "原产地"
    sheet["B13"] = "贴纸"
    sheet["C13"] = "个"
    sheet["D13"] = 100
    sheet["H13"] = 0.0013
    sheet["J13"] = 0.5013
    sheet["Q13"] = "4821900000纸或纸板的其他各种标签"
    sheet["R13"] = "0:品牌类型：0|1:出口享惠情况：1"
    sheet["S13"] = "中国"
    for row in sheet.iter_rows(min_row=5, max_row=13, min_col=1, max_col=19):
        for cell in row:
            cell.border = _thin()


def _auxiliary(sheet) -> None:
    sheet["B1"] = "报关单号"
    sheet["C1"] = "报关日期"
    sheet["E1"] = "恒信编号"
    sheet["K1"] = "SAP编号"
    sheet["H1"] = "商品编码"
    sheet["M1"] = "商品名称"
    sheet["B2"] = "SHOULD-NOT-MAP"
    sheet["H2"] = "9999999999"
    sheet["M2"] = "内部料号壳体"
    for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=13):
        for cell in row:
            cell.border = _thin()


def _values(item) -> dict[str, str]:
    return {name: field.value for name, field in item.fields.items() if field.value}


def test_hengxin_draft_maps_core_goods_columns() -> None:
    document = parse_excel(
        _workbook({"一般贸易出口": _draft}),
        file_id="hx",
        filename="hengxin-draft.xlsx",
    )
    draft = document.sheets[0]
    assert draft.role == "draft"
    items = map_sheet_goods(draft, document)
    assert len(items) == 1
    values = _values(items[0])
    assert values["gno"] == "1"
    assert values["codeTs"] == "9111900000"
    assert values["gname"] == "表壳配件/壳体"
    assert values["gmodel"] == "不锈钢/无中文品牌、无外文品牌/无型号"
    assert items[0].fields["gmodel"].status.value == "needs_review"
    assert values["gqty"] == "150"
    assert values["gunit"] == "只"
    assert values["declPrice"] == "98"
    assert values["declTotal"] == "14700"
    assert values["tradeCurr"] == "港币"
    assert values["cusOriginCountry"] == "中国"
    assert values["destinationCountry"] == "中国香港"
    assert values["districtCode"] == "惠州其他"
    assert values["customNetWt"] == "5.74"
    assert "qty1" not in values
    assert all(field.evidence for field in items[0].fields.values())


def test_ten_digit_customs_code_header_maps_hs() -> None:
    def ten_digit(sheet) -> None:
        sheet["A1"] = "PACKING LIST"
        sheet["A3"] = "物料名称"
        sheet["B3"] = "海关十位编码"
        sheet["C3"] = "出货数量"
        sheet["A4"] = "贴纸"
        sheet["B4"] = "4821900000纸或纸板的其他各种标签"
        sheet["C4"] = 100
        for row in sheet.iter_rows(min_row=1, max_row=4, min_col=1, max_col=3):
            for cell in row:
                cell.border = _thin()

    document = parse_excel(
        _workbook({"总箱单": ten_digit}),
        file_id="hs10",
        filename="hs10.xlsx",
    )
    items = map_document_goods(document)
    assert items[0].value_of("codeTs") == "4821900000"
    assert items[0].value_of("gname") == "贴纸"


def test_same_role_pages_concat_instead_of_row_align() -> None:
    """两页 draft 项号 1–2 / 3–4 接成 4 行，不把第 2 页第 1 件补进第 1 件（#23）。"""
    from docparse.adapters.parsers.layout import split_sheet
    from docparse.domain.ir import Cell, DocumentIR, Sheet
    from docparse.extraction.sheet_role import classify_sheet

    def _page(name: str, start: int, count: int) -> Sheet:
        cells = [
            Cell(address="A1", value="中华人民共和国海关出口货物报关单", row=1, column=1),
            Cell(address="A2", value="项号", row=2, column=1),
            Cell(address="B2", value="商品编号", row=2, column=2),
            Cell(address="C2", value="商品名称及规格型号", row=2, column=3),
            Cell(address="D2", value="数量", row=2, column=4),
        ]
        for offset in range(count):
            gno = start + offset
            row = 3 + offset
            cells.extend(
                [
                    Cell(address=f"A{row}", value=str(gno), row=row, column=1),
                    Cell(address=f"B{row}", value=f"190531000{gno}", row=row, column=2),
                    Cell(address=f"C{row}", value=f"商品{gno}", row=row, column=3),
                    Cell(address=f"D{row}", value=str(gno * 10), row=row, column=4),
                ]
            )
        sheet = split_sheet(Sheet(name=name, cells=cells))
        return classify_sheet(sheet, filename="scan.pdf")

    document = DocumentIR(
        document_id="pages",
        file_id="f",
        filename="scan.pdf",
        media_type="application/pdf",
        sheets=[_page("1", 1, 2), _page("2", 3, 2)],
    )
    items = map_document_goods(document)
    assert [item.value_of("gno") for item in items] == ["1", "2", "3", "4"]
    assert [item.value_of("gname") for item in items] == ["商品1", "商品2", "商品3", "商品4"]


def test_packing_fills_gross_when_qty_aligns() -> None:
    document = parse_excel(
        _workbook({"一般贸易出口": _draft, "装箱单": _packing}),
        file_id="mix",
        filename="hengxin.xlsx",
    )
    items = map_document_goods(document)
    assert len(items) == 1
    values = _values(items[0])
    assert items[0].source_role == "draft"
    assert values["gqty"] == "150"
    assert values["customNetWt"] == "5.74"
    assert values["customGrossWet"] == "7.35"
    assert items[0].fields["customGrossWet"].evidence[0].cell.startswith("装箱单!")


def test_skip_fill_knob_blocks_configured_field() -> None:
    schema = load_schema().model_copy(deep=True)
    schema.goods_master.skip_fill = ["customGrossWet"]
    document = parse_excel(
        _workbook({"一般贸易出口": _draft, "装箱单": _packing}),
        file_id="skip",
        filename="skip.xlsx",
    )
    items = map_document_goods(document, schema)
    assert items[0].value_of("gqty") == "150"
    assert items[0].value_of("customGrossWet") is None


def test_qty_mismatch_does_not_fill_qty_but_fills_gross() -> None:
    def other_qty(sheet) -> None:
        _packing(sheet)
        sheet["E12"] = 500

    document = parse_excel(
        _workbook({"一般贸易出口": _draft, "装箱单": other_qty}),
        file_id="mis",
        filename="mismatch-qty.xlsx",
    )
    items = map_document_goods(document)
    assert len(items) == 1
    assert items[0].value_of("gqty") == "150"
    assert items[0].value_of("customNetWt") == "5.74"
    assert items[0].value_of("customGrossWet") == "7.35"


def test_gross_below_net_is_not_filled() -> None:
    def zero_gross(sheet) -> None:
        _packing(sheet)
        sheet["I12"] = 0

    document = parse_excel(
        _workbook({"一般贸易出口": _draft, "装箱单": zero_gross}),
        file_id="zero",
        filename="zero-gross.xlsx",
    )
    items = map_document_goods(document)
    assert items[0].value_of("gqty") == "150"
    assert items[0].value_of("customNetWt") == "5.74"
    assert items[0].value_of("customGrossWet") is None


def test_unit_mismatch_blocks_qty_chain_but_not_gross() -> None:
    def kg_packing(sheet) -> None:
        _packing(sheet)
        sheet["E12"] = None
        sheet["F11"] = "Unit(单位）"
        sheet["F12"] = "千克"

    def pc_packing(sheet) -> None:
        _packing(sheet)
        sheet["H12"] = None

    document = parse_excel(
        _workbook({"一般贸易出口": _draft, "装箱单": pc_packing, "箱单2": kg_packing}),
        file_id="unit",
        filename="unit.xlsx",
    )
    items = map_document_goods(document)
    item = items[0]
    assert item.value_of("gunit") == "只"
    assert item.value_of("gqty") == "150"
    assert item.value_of("declPrice") == "98"
    assert item.value_of("declTotal") == "14700"
    assert item.value_of("customGrossWet") == "7.35"


def test_kg_row_uses_net_weight_not_other_qty() -> None:
    def kg_draft(sheet) -> None:
        _draft(sheet)
        sheet["E18"] = None
        sheet["F18"] = "千克"
        sheet["G18"] = 5.74
        sheet["H18"] = 98
        sheet["I18"] = 562.52

    def pcs_packing(sheet) -> None:
        _packing(sheet)
        sheet["E12"] = 500
        sheet["I12"] = 7.35

    document = parse_excel(
        _workbook({"一般贸易出口": kg_draft, "装箱单": pcs_packing}),
        file_id="kg",
        filename="kg.xlsx",
    )
    items = map_document_goods(document)
    assert items[0].value_of("gunit") == "千克"
    assert items[0].value_of("gqty") is None
    assert items[0].value_of("customNetWt") == "5.74"
    assert items[0].value_of("customGrossWet") == "7.35"


def test_kg_row_fills_qty_only_when_close_to_net() -> None:
    def kg_draft(sheet) -> None:
        _draft(sheet)
        sheet["E18"] = None
        sheet["F18"] = "千克"
        sheet["G18"] = 5.74
        sheet["H18"] = 98
        sheet["I18"] = 562.52

    def kg_packing(sheet) -> None:
        _packing(sheet)
        sheet["F11"] = "Unit(单位）"
        sheet["F12"] = "千克"
        sheet["E12"] = 500

    def kg_invoice(sheet) -> None:
        _invoice(sheet)
        sheet["D17"] = 5.74
        sheet["F17"] = 98
        sheet["G17"] = 562.52

    document = parse_excel(
        _workbook({"一般贸易出口": kg_draft, "装箱单": kg_packing, "发票": kg_invoice}),
        file_id="kgfill",
        filename="kgfill.xlsx",
    )
    items = map_document_goods(document)
    item = items[0]
    assert item.value_of("gunit") == "千克"
    assert item.value_of("gqty") == "5.74"
    assert item.fields["gqty"].evidence[0].cell.startswith("发票!")
    assert item.value_of("customGrossWet") == "7.35"
    assert item.fields["customGrossWet"].evidence[0].cell.startswith("装箱单!")


def test_kg_row_same_table_qty_takes_net_not_pcs() -> None:
    """同行同时有重量、数量PCS、计价单位=千克：成交数量取净重。"""

    def kg_pcs_draft(sheet) -> None:
        _draft(sheet)
        sheet["E18"] = 5000
        sheet["F18"] = "千克"
        sheet["G18"] = 205
        sheet["H18"] = 3810
        sheet["I18"] = 781050

    document = parse_excel(
        _workbook({"一般贸易出口": kg_pcs_draft}),
        file_id="kgpcs",
        filename="kg-pcs.xlsx",
    )
    items = map_document_goods(document)
    item = items[0]
    assert item.value_of("gunit") == "千克"
    assert item.value_of("customNetWt") == "205"
    assert item.value_of("gqty") == "205"
    assert "重量KG" in (item.fields["gqty"].evidence[0].quote or "")
    assert item.value_of("declPrice") == "3810"
    assert item.value_of("declTotal") == "781050"


def test_piece_unit_keeps_pcs_qty() -> None:
    document = parse_excel(
        _workbook({"一般贸易出口": _draft}),
        file_id="pcs",
        filename="pcs.xlsx",
    )
    item = map_document_goods(document)[0]
    assert item.value_of("gunit") == "只"
    assert item.value_of("gqty") == "150"
    assert item.value_of("customNetWt") == "5.74"


def test_kg_row_without_net_keeps_qty() -> None:
    def kg_qty_only(sheet) -> None:
        _draft(sheet)
        sheet["E18"] = 17432
        sheet["F18"] = "千克"
        sheet["G18"] = None
        sheet["H18"] = 33.9028
        sheet["I18"] = 590993

    document = parse_excel(
        _workbook({"一般贸易出口": kg_qty_only}),
        file_id="kgonly",
        filename="kg-qty-only.xlsx",
    )
    item = map_document_goods(document)[0]
    assert item.value_of("gunit") == "千克"
    assert item.value_of("gqty") == "17432"
    assert item.value_of("customNetWt") is None


def test_invoice_qty_fills_when_price_implies_same_qty() -> None:
    def priced(sheet) -> None:
        _draft(sheet)
        sheet["E18"] = None
        sheet["H18"] = 98
        sheet["I18"] = 14700

    def counted(sheet) -> None:
        _packing(sheet)
        sheet["E12"] = 150
        sheet["H12"] = None
        sheet["I12"] = None

    document = parse_excel(
        _workbook({"一般贸易出口": priced, "装箱单": counted}),
        file_id="swap",
        filename="price-then-qty.xlsx",
    )
    items = map_document_goods(document)
    assert items[0].value_of("declPrice") == "98"
    assert items[0].value_of("declTotal") == "14700"
    assert items[0].value_of("gqty") == "150"


def test_weight_without_gross_leaves_gross_empty() -> None:
    document = parse_excel(
        _workbook({"一般贸易出口": _draft}),
        file_id="hx",
        filename="draft-only.xlsx",
    )
    items = map_document_goods(document)
    values = _values(items[0])
    assert values["customNetWt"] == "5.74"
    assert items[0].value_of("customGrossWet") is None
    assert "net_as_gross" not in items[0].review_reasons


def test_auxiliary_is_not_master_or_supplement() -> None:
    document = parse_excel(
        _workbook({"一般贸易出口": _draft, "Sheet3": _auxiliary}),
        file_id="aux",
        filename="aux.xlsx",
    )
    aux = next(sheet for sheet in document.sheets if sheet.name == "Sheet3")
    assert aux.consume == "exclude"
    assert map_sheet_goods(aux, document) == []
    items = map_document_goods(document)
    assert len(items) == 1
    assert items[0].value_of("codeTs") == "9111900000"
    assert all(item.source_role != "auxiliary" for item in items)


def test_guoguang_packing_is_master_invoice_fills_price() -> None:
    document = parse_excel(
        _workbook({"总箱单": _guoguang_packing, "发票": _invoice}),
        file_id="gg",
        filename="guoguang.xlsx",
    )
    packing = next(sheet for sheet in document.sheets if sheet.role == "packing")
    invoice = next(sheet for sheet in document.sheets if sheet.role == "invoice")
    pack_items = map_sheet_goods(packing, document)
    inv_items = map_sheet_goods(invoice, document)
    assert pack_items[0].master_score > inv_items[0].master_score
    items = map_document_goods(document)
    assert len(items) == 1
    values = _values(items[0])
    assert items[0].source_role == "packing"
    assert values["gname"] == "贴纸"
    assert values["codeTs"] == "4821900000"
    assert values["gqty"] == "100"
    assert values["declPrice"] == "0.0181"
    assert values["declTotal"] == "1.81"
    assert items[0].fields["declPrice"].evidence[0].cell.startswith("发票!")


def test_same_name_rows_match_by_qty() -> None:
    def two_caps(sheet) -> None:
        _guoguang_packing(sheet)
        sheet["B14"] = "电容"
        sheet["C14"] = "个"
        sheet["D14"] = 10000
        sheet["H14"] = 0.45
        sheet["Q14"] = "8532241000"
        sheet["R14"] = "瓷介质"
        sheet["S14"] = "中国"
        sheet["B15"] = "电容"
        sheet["C15"] = "个"
        sheet["D15"] = 16000
        sheet["H15"] = 0.57
        sheet["Q15"] = "8532241000"
        sheet["R15"] = "瓷介质"
        sheet["S15"] = "中国"
        for row in sheet.iter_rows(min_row=14, max_row=15, min_col=1, max_col=19):
            for cell in row:
                cell.border = _thin()

    def two_prices(sheet) -> None:
        _invoice(sheet)
        sheet["B18"] = "电容"
        sheet["D18"] = 10000
        sheet["E18"] = "USD"
        sheet["F18"] = 0.00031
        sheet["G18"] = 3.1
        sheet["B19"] = "电容"
        sheet["D19"] = 16000
        sheet["E19"] = "USD"
        sheet["F19"] = 0.00346
        sheet["G19"] = 55.36
        for row in sheet.iter_rows(min_row=18, max_row=19, min_col=1, max_col=7):
            for cell in row:
                cell.border = _thin()

    document = parse_excel(
        _workbook({"总箱单": two_caps, "发票": two_prices}),
        file_id="dup",
        filename="dup.xlsx",
    )
    items = map_document_goods(document)
    caps = [item for item in items if item.value_of("gname") == "电容"]
    assert len(caps) == 2
    by_qty = {item.value_of("gqty"): item.value_of("declPrice") for item in caps}
    assert by_qty == {"10000": "0.00031", "16000": "0.00346"}
    assert all(item.source_kind == "primary" for item in caps)


def test_unmatched_packing_row_is_supplement() -> None:
    def extra_packing(sheet) -> None:
        _packing(sheet)
        sheet["C13"] = "配件"
        sheet["E13"] = 10
        sheet["I13"] = 1.2
        for row in sheet.iter_rows(min_row=13, max_row=13, min_col=1, max_col=10):
            for cell in row:
                cell.border = _thin()

    document = parse_excel(
        _workbook({"一般贸易出口": _draft, "装箱单": extra_packing}),
        file_id="extra",
        filename="extra.xlsx",
    )
    items = map_document_goods(document)
    assert len(items) == 1
    assert all(item.source_kind == "primary" for item in items)
    assert all(item.value_of("gname") != "配件" for item in items)


@pytest.mark.skipif(not REAL_HENGXIN.exists(), reason="本地恒信样本不在 CI")
def test_hengxin_sample_goods() -> None:
    document = parse_excel(
        REAL_HENGXIN.read_bytes(),
        file_id="hx-real",
        filename=REAL_HENGXIN.name,
    )
    items = map_document_goods(document)
    first = items[0]
    values = _values(first)
    assert first.source_role == "draft"
    assert values["gno"] == "1"
    assert values["codeTs"] == "9111900000"
    assert values["gname"] == "表壳配件/壳体"
    assert "不锈钢" in (values["gmodel"] or "")
    assert values["customNetWt"] == "5.74"
    assert all(item.source_kind == "primary" for item in items)
    assert all(item.source_role != "auxiliary" for item in items)
    if len(items) >= 2:
        second = items[1]
        assert second.value_of("gunit") == "千克"
        assert second.value_of("gqty") != "500"
        assert second.value_of("gqty") == second.value_of("customNetWt")
        assert second.fields["gqty"].evidence[0].cell.startswith("发票!")
        assert second.value_of("customGrossWet") == "10.38"
        assert second.fields["customGrossWet"].evidence[0].cell.startswith("裝箱單!")
    assert first.value_of("customGrossWet") == "7.35"
    # G.W.=0 的拼箱行（净重 0.05 / 0.03）：0 < 净重，毛重不补
    by_gno = {item.value_of("gno"): item for item in items}
    assert by_gno["15"].value_of("customNetWt") == "0.05"
    assert by_gno["15"].value_of("customGrossWet") is None
    assert by_gno["26"].value_of("customNetWt") == "0.03"
    assert by_gno["26"].value_of("customGrossWet") is None
    assert all(field.evidence for field in first.fields.values())


@pytest.mark.skipif(not REAL_GUOGUANG.exists(), reason="本地国光样本不在 CI")
def test_guoguang_sample_goods() -> None:
    document = parse_excel(
        REAL_GUOGUANG.read_bytes(),
        file_id="gg-real",
        filename=REAL_GUOGUANG.name,
    )
    items = map_document_goods(document)
    first = items[0]
    values = _values(first)
    assert first.source_role == "packing"
    assert values["gname"] == "贴纸"
    assert values["codeTs"] == "4821900000"
    assert values["declPrice"] == "0.0181"
    assert all(item.source_kind == "primary" for item in items)
    assert all(field.evidence for field in first.fields.values())


def _continuation_two_row(sheet) -> None:
    """两行一项：行 2 只有申报要素，落在品名列。"""
    sheet["A1"] = "中华人民共和国海关出口货物报关单"
    headers = [
        ("A3", "项号"),
        ("B3", "商品编号"),
        ("C3", "商品名称及规格型号"),
        ("D3", "数量"),
        ("E3", "单位"),
        ("F3", "单价"),
        ("G3", "总价"),
        ("H3", "币制"),
    ]
    for address, text in headers:
        sheet[address] = text
    sheet["A4"] = 1
    sheet["B4"] = "8708100000"
    sheet["C4"] = "前保险杠下部装饰板"
    sheet["D4"] = 1
    sheet["E4"] = "个"
    sheet["F4"] = 86.03
    sheet["G4"] = 86.03
    sheet["H4"] = "CNY"
    sheet["C5"] = "境内自主品牌|不享惠|吉利EC7 1.5L/1.8L等小轿车通用|吉利牌|零部件编号6010182800"
    sheet["A6"] = 2
    sheet["B6"] = "8708999900"
    sheet["C6"] = "喷水壶带洗涤电机总成"
    sheet["D6"] = 10
    sheet["E6"] = "个"
    sheet["F6"] = 89.35
    sheet["G6"] = 893.5
    sheet["H6"] = "CNY"
    sheet["C7"] = "境内自主品牌|不享惠|吉利EC7 1.5L/1.8L等小轿车通用|吉利牌|零部件编号6608056397"
    for row in sheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=8):
        for cell in row:
            cell.border = _thin()


def _continuation_three_row(sheet) -> None:
    """三行一项：单价/总价/币制竖排，数量及单位竖排两行。"""
    sheet["A1"] = "中华人民共和国海关出口货物报关单"
    headers = [
        ("A3", "项号"),
        ("B3", "商品编号"),
        ("C3", "商品名称及规格型号"),
        ("D3", "数量及单位"),
        ("E3", "单价/总价/币制"),
        ("F3", "原产国"),
        ("G3", "最终目的国"),
        ("H3", "境内货源地"),
        ("I3", "征免"),
    ]
    for address, text in headers:
        sheet[address] = text
    sheet["A4"] = 1
    sheet["B4"] = "4901990000"
    sheet["C4"] = "外文书籍"
    sheet["D4"] = 17432
    sheet["E4"] = 33.9028
    sheet["F4"] = "中国"
    sheet["G4"] = "美国"
    sheet["H4"] = "(44199/)东莞/"
    sheet["I4"] = "照章征税"
    sheet["C5"] = "0|0|其他书籍|成册|||无品牌"
    sheet["D5"] = 0
    sheet["E5"] = 590993
    sheet["F5"] = "(CHN)"
    sheet["G5"] = "(USA)"
    sheet["I5"] = "(1)"
    sheet["A6"] = 0
    sheet["D6"] = "千克"
    sheet["E6"] = "人民币"
    for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=9):
        for cell in row:
            cell.border = _thin()


def _total_row_draft(sheet) -> None:
    """主行后跟合计行：合计不能成商品，也不能并进最后一项。"""
    _draft(sheet)
    sheet["A19"] = "合计："
    sheet["E19"] = 150
    sheet["G19"] = 5.74
    sheet["I19"] = 14700
    for row in sheet.iter_rows(min_row=19, max_row=19, min_col=1, max_col=13):
        for cell in row:
            cell.border = _thin()


def test_two_row_item_merges_declaration_element() -> None:
    document = parse_excel(
        _workbook({"一般贸易出口": _continuation_two_row}),
        file_id="two",
        filename="two-row.xlsx",
    )
    items = map_document_goods(document)
    assert len(items) == 2
    first = _values(items[0])
    assert first["gno"] == "1"
    assert first["codeTs"] == "8708100000"
    assert first["gname"] == "前保险杠下部装饰板"
    assert first["gmodel"] == (
        "境内自主品牌|不享惠|吉利EC7 1.5L/1.8L等小轿车通用|吉利牌|零部件编号6010182800"
    )
    assert first["gqty"] == "1"
    second = _values(items[1])
    assert second["gno"] == "2"
    assert second["gname"] == "喷水壶带洗涤电机总成"
    assert "零部件编号6608056397" in (second["gmodel"] or "")
    assert all("|" in (item.value_of("gmodel") or "") for item in items)


def test_three_row_item_fills_stacked_price_and_unit() -> None:
    document = parse_excel(
        _workbook({"报关预录入单": _continuation_three_row}),
        file_id="three",
        filename="three-row.xlsx",
    )
    items = map_document_goods(document)
    assert len(items) == 1
    values = _values(items[0])
    assert values["codeTs"] == "4901990000"
    assert values["gname"] == "外文书籍"
    assert values["gqty"] == "17432"
    assert values["gunit"] == "千克"
    assert values["declPrice"] == "33.9028"
    assert values["declTotal"] == "590993"
    assert values["tradeCurr"] == "人民币"
    assert values["gmodel"] == "0|0|其他书籍|成册|||无品牌"
    assert values["cusOriginCountry"] == "中国"
    assert values["destinationCountry"] == "美国"
    assert values["districtCode"] == "(44199/)东莞/"
    assert values["dutyMode"] == "照章征税"
    assert items[0].value_of("gno") == "1"


def test_total_row_is_dropped_not_merged() -> None:
    document = parse_excel(
        _workbook({"一般贸易出口": _total_row_draft}),
        file_id="tot",
        filename="total-row.xlsx",
    )
    items = map_document_goods(document)
    assert len(items) == 1
    values = _values(items[0])
    assert values["gno"] == "1"
    assert values["gname"] == "表壳配件/壳体"
    assert values["gqty"] == "150"
    assert values["declTotal"] == "14700"
    assert all("合计" not in (item.value_of("gno") or "") for item in items)
    assert all("合计" not in (item.value_of("gname") or "") for item in items)


def test_total_label_outside_mapped_column_is_still_dropped() -> None:
    """合计落在无表头中间列（MXY 装箱单），物理行扫描也要拦住。"""

    def packing_total(sheet) -> None:
        sheet["A1"] = "PACKING LIST"
        sheet["A3"] = "序号"
        sheet["B3"] = "DESCRIPTION"
        sheet["D3"] = "Ship Q'ty (数量/PCS)"
        sheet["E3"] = "N.W. (Kg) (净重)"
        sheet["F3"] = "G.W .(Kg) (毛重)"
        sheet["A4"] = 1
        sheet["B4"] = "UPS电源系统"
        sheet["D4"] = 1
        sheet["E4"] = 733
        sheet["F4"] = 813
        sheet["C5"] = "合计"
        sheet["D5"] = 1
        sheet["E5"] = 733
        sheet["F5"] = 813
        for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=6):
            for cell in row:
                cell.border = _thin()

    document = parse_excel(
        _workbook({"装箱单": packing_total}),
        file_id="ptot",
        filename="packing-total.xlsx",
    )
    items = map_document_goods(document)
    assert len(items) == 1
    assert items[0].value_of("gname") == "UPS电源系统"
    assert items[0].value_of("customGrossWet") == "813"


def test_orphan_continuation_without_master_is_dropped() -> None:
    def orphan(sheet) -> None:
        sheet["A1"] = "中华人民共和国海关出口货物报关单"
        sheet["A3"] = "项号"
        sheet["B3"] = "商品编号"
        sheet["C3"] = "商品名称及规格型号"
        sheet["C4"] = "境内自主品牌|不享惠|吉利EC7|吉利牌|零部件"
        sheet["A5"] = 1
        sheet["B5"] = "8708100000"
        sheet["C5"] = "前保险杠"
        for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3):
            for cell in row:
                cell.border = _thin()

    document = parse_excel(
        _workbook({"一般贸易出口": orphan}),
        file_id="orphan",
        filename="orphan.xlsx",
    )
    items = map_document_goods(document)
    assert len(items) == 1
    assert items[0].value_of("gname") == "前保险杠"
    assert items[0].value_of("gmodel") is None


@pytest.mark.skipif(not REAL_GSRUA.exists(), reason="本地 GSRUA 样本不在 CI")
def test_gsrua_sample_merges_two_row_items() -> None:
    document = parse_excel(
        REAL_GSRUA.read_bytes(),
        file_id="gsrua",
        filename=REAL_GSRUA.name,
    )
    items = map_document_goods(document)
    assert len(items) == 50
    assert all(item.value_of("codeTs") for item in items)
    assert all((item.value_of("gmodel") or "").count("|") >= 2 for item in items)
    first = _values(items[0])
    assert first["gno"] == "1"
    assert first["codeTs"] == "8708100000"
    assert first["gname"] == "前保险杠下部装饰板"
    assert first["gmodel"].startswith("境内自主品牌|不享惠|吉利EC7")
    assert all("合计" not in (item.value_of("gno") or "") for item in items)


@pytest.mark.skipif(not REAL_DONNELLEY.exists(), reason="本地当纳利样本不在 CI")
def test_donnelley_sample_merges_three_row_item() -> None:
    document = parse_excel(
        REAL_DONNELLEY.read_bytes(),
        file_id="dnn",
        filename=REAL_DONNELLEY.name,
    )
    items = map_document_goods(document)
    assert len(items) == 1
    values = _values(items[0])
    assert values["codeTs"] == "4901990000"
    assert values["gname"] == "外文书籍"
    assert values["gqty"] == "17432"
    assert values["gunit"] == "千克"
    assert values["declPrice"] == "33.9028"
    assert values["declTotal"] == "590993"
    assert values["tradeCurr"] == "人民币"
    assert values["gmodel"] == "0|0|其他书籍|成册|||无品牌"
    assert values["cusOriginCountry"] == "中国"
    assert values["destinationCountry"] == "美国"
    assert values["districtCode"] == "(44199/)东莞/"
    assert values["dutyMode"] == "照章征税"


@pytest.mark.skipif(not REAL_BLU_IN.exists(), reason="本地进料 BLU 样本不在 CI")
def test_blu_in_kg_qty_takes_net_not_pcs() -> None:
    document = parse_excel(
        REAL_BLU_IN.read_bytes(),
        file_id="blu-in",
        filename=REAL_BLU_IN.name,
    )
    items = map_document_goods(document)
    assert len(items) == 49
    first = items[0]
    assert first.value_of("gunit") == "千克"
    assert first.value_of("customNetWt") == "205"
    assert first.value_of("gqty") == "205"
    assert first.value_of("declPrice") == "3810"
    assert first.value_of("declTotal") == "781050"
    for item in items:
        assert item.value_of("gunit") == "千克"
        net = item.value_of("customNetWt")
        qty = item.value_of("gqty")
        assert net and qty
        assert qty == net
        price = float(item.value_of("declPrice") or "0")
        total = float(item.value_of("declTotal") or "0")
        if price:
            assert abs(price * float(qty) - total) <= 0.05 or abs(
                price * float(qty) - total
            ) <= 0.005 * max(abs(total), 1.0)


@pytest.mark.skipif(not REAL_MXY.exists(), reason="本地 MXY 样本不在 CI")
def test_mxy_sample_drops_total_and_keeps_seven() -> None:
    document = parse_excel(
        REAL_MXY.read_bytes(),
        file_id="mxy",
        filename=REAL_MXY.name,
    )
    items = map_document_goods(document)
    assert len(items) == 7
    assert [item.value_of("gno") for item in items] == [str(i) for i in range(1, 8)]
    assert all("合计" not in (item.value_of("gname") or "") for item in items)


@pytest.mark.skipif(not REAL_DUOKE.exists(), reason="本地多科样本不在 CI")
def test_duoke_sample_keeps_four_items() -> None:
    document = parse_excel(
        REAL_DUOKE.read_bytes(),
        file_id="dk",
        filename=REAL_DUOKE.name,
    )
    items = map_document_goods(document)
    assert len(items) == 4
    assert [item.value_of("gno") for item in items] == ["1", "2", "3", "4"]
