from __future__ import annotations

import io
from pathlib import Path

import pytest

pytest.importorskip("openpyxl")

from openpyxl import Workbook
from openpyxl.styles import Border, Side

from docparse.adapters.parsers.excel import parse_excel
from docparse.extraction.assemble import assemble_declaration, declaration_payload

_DEMO = Path("/Users/baldwin/Desktop/taizhou/AI识别Demo")
REAL_HENGXIN = _DEMO / "（恒信）一般贸易草单HDX260251BLU.xlsx"
REAL_GUOGUANG = _DEMO / "（国光）箱单发票合同26VN0502-1.xlsx"


def _thin() -> Border:
    line = Side(style="thin")
    return Border(left=line, right=line, top=line, bottom=line)


def _workbook(builders: dict) -> bytes:
    book = Workbook()
    first = True
    for title, fill in builders.items():
        sheet = book.active if first else book.create_sheet(title)
        if first:
            sheet.title = title
            first = False
        fill(sheet)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _draft(sheet) -> None:
    sheet["A1"] = "中华人民共和国海关出口货物报关单"
    sheet.merge_cells("A3:D3")
    sheet["A3"] = "境内发货人"
    sheet.merge_cells("A4:D4")
    sheet["A4"] = "惠州市恒德信精密科技有限公司441394164D"
    sheet["E3"] = "出境关别"
    sheet["E4"] = "莲塘口岸"
    sheet.merge_cells("A5:D5")
    sheet["A5"] = "境外收货人"
    sheet.merge_cells("A6:D6")
    sheet["A6"] = "BLU PRECISION LIMITED"
    sheet.merge_cells("A7:D7")
    sheet["A7"] = "生产销售单位"
    sheet.merge_cells("A8:D8")
    sheet["A8"] = "惠州市恒德信精密科技有限公司"
    sheet["E5"] = "运输方式"
    sheet["E6"] = "公路运输"
    sheet["E7"] = "监管方式"
    sheet["E8"] = "一般贸易"
    sheet.merge_cells("A9:D9")
    sheet["A9"] = "合同协议号"
    sheet.merge_cells("A10:D10")
    sheet["A10"] = "HDX2026-251"
    sheet.merge_cells("A11:C11")
    sheet["A11"] = "包装种类"
    sheet["D11"] = "件数"
    sheet["E11"] = "毛重（千克）"
    sheet["F11"] = "净重（千克）"
    sheet.merge_cells("G11:H11")
    sheet["G11"] = "成交方式"
    sheet.merge_cells("A12:C12")
    sheet["A12"] = "其它"
    sheet["D12"] = 40
    sheet["E12"] = 296.46
    sheet["F12"] = 218.375
    sheet.merge_cells("G12:H12")
    sheet["G12"] = "FOB"
    headers = [
        ("A17", "项号"),
        ("B17", "商品编号"),
        ("C17", "商品名称及规格型号"),
        ("D17", "申报要素"),
        ("E17", "数量PCS"),
        ("F17", "计价单位"),
        ("G17", "重量KG"),
        ("H17", "单价"),
        ("I17", "总价"),
        ("J17", "币制"),
        ("K17", "原产国（地区）"),
        ("L17", "最终目的国（地区）"),
        ("M17", "境内货源地"),
    ]
    for address, text in headers:
        sheet[address] = text
    sheet["A18"] = 1
    sheet["B18"] = "9111900000"
    sheet["C18"] = "表壳配件/壳体"
    sheet["D18"] = "不锈钢/无中文品牌、无外文品牌/无型号"
    sheet["E18"] = 150
    sheet["F18"] = "只"
    sheet["G18"] = 5.74
    sheet["H18"] = 98
    sheet["I18"] = 14700
    sheet["J18"] = "港币"
    sheet["K18"] = "中国"
    sheet["L18"] = "中国香港"
    sheet["M18"] = "惠州其他"
    for row in sheet.iter_rows(min_row=1, max_row=18, min_col=1, max_col=13):
        for cell in row:
            cell.border = _thin()


def _packing(sheet) -> None:
    sheet["A1"] = "PACKING LIST"
    sheet["I4"] = "Invoice No.﹕"
    sheet["J4"] = "HDX260251"
    sheet["A7"] = "Bill To : BLU PRECISION LIMITED"
    sheet["C11"] = "Description (货物名称)"
    sheet["E11"] = "Ship Q'ty (数量/PCS)"
    sheet["G11"] = "Packing (箱)"
    sheet["H11"] = "N.W. (Kg) (净重)"
    sheet["I11"] = "G.W .(Kg) (毛重)"
    sheet["C12"] = "表壳配件/壳体"
    sheet["E12"] = 150
    sheet["G12"] = 40
    sheet["H12"] = 5.74
    sheet["I12"] = 7.35
    for row in sheet.iter_rows(min_row=1, max_row=12, min_col=1, max_col=10):
        for cell in row:
            cell.border = _thin()


def _mismatch_packing(sheet) -> None:
    _packing(sheet)
    sheet["A3"] = "件数"
    sheet["B3"] = 99


def _net_only_packing(sheet) -> None:
    sheet["A1"] = "PACKING LIST"
    sheet["A3"] = "净重"
    sheet["B3"] = 2825.47
    sheet["A6"] = "合同号CONTRACT NO.:"
    sheet["B6"] = "26VN0502"
    sheet["A8"] = "卖方SELLER"
    sheet["A9"] = "GUOGUANG ELECTRIC COMPANY LIMITED."
    sheet["F8"] = "买方BUYER"
    sheet["F9"] = "GUOGUANG ACOUSTICS (VIETNAM) COMPANY LIMITED"
    sheet["B12"] = "物料名称"
    sheet["C12"] = "单位"
    sheet["D12"] = "出货数量"
    sheet["H12"] = "总净重 NW"
    sheet["Q12"] = "HS CODE海关编码"
    sheet["B13"] = "贴纸"
    sheet["C13"] = "个"
    sheet["D13"] = 100
    sheet["H13"] = 0.0013
    sheet["Q13"] = "4821900000纸或纸板的其他各种标签"
    for row in sheet.iter_rows(min_row=1, max_row=13, min_col=1, max_col=19):
        for cell in row:
            cell.border = _thin()


def _invoice(sheet) -> None:
    sheet["A1"] = "INVOICE"
    sheet["F6"] = "发票号INVOICE NO.:"
    sheet["G6"] = "26VN0502-1"
    sheet["A8"] = "买方BUYER"
    sheet["A9"] = "GUOGUANG ACOUSTICS (VIETNAM) COMPANY LIMITED"
    sheet["G7"] = "合同号CONTRACT NO.:"
    sheet["H7"] = "26VN0502"
    sheet["B16"] = "物料名称"
    sheet["D16"] = "出货数量"
    sheet["E16"] = "币别"
    sheet["F16"] = "单价"
    sheet["G16"] = "汇总价"
    sheet["B17"] = "贴纸"
    sheet["D17"] = 100
    sheet["E17"] = "USD"
    sheet["F17"] = 0.0181
    sheet["G17"] = 1.81
    for row in sheet.iter_rows(min_row=1, max_row=17, min_col=1, max_col=8):
        for cell in row:
            cell.border = _thin()


def _auxiliary(sheet) -> None:
    sheet["B1"] = "报关单号"
    sheet["C1"] = "报关日期"
    sheet["E1"] = "恒信编号"
    sheet["K1"] = "SAP编号"
    sheet["H1"] = "商品编码"
    sheet["M1"] = "商品名称"
    sheet["A3"] = "合同协议号"
    sheet["A4"] = "SHOULD-NOT-ASSEMBLE"
    sheet["H2"] = "9999999999"
    sheet["M2"] = "内部料号壳体"
    for row in sheet.iter_rows(min_row=1, max_row=4, min_col=1, max_col=13):
        for cell in row:
            cell.border = _thin()


def _parse(builders: dict, filename: str):
    return parse_excel(_workbook(builders), file_id="asm", filename=filename)


def test_draft_is_copied_and_codes_are_looked_up() -> None:
    document = _parse(
        {"一般贸易出口": _draft, "装箱单": _packing, "Sheet3": _auxiliary},
        "hengxin.xlsx",
    )
    declaration = assemble_declaration(
        document,
        agent={"agentCode": "4403180867", "agentName": "深圳市泰洲物流有限公司"},
    )
    payload = declaration_payload(declaration)

    assert declaration.has_draft is True
    assert declaration.source_roles == ["draft", "packing"]
    assert payload["contrNo"] == "HDX2026-251"
    assert payload["consignorEname"] == "BLU PRECISION LIMITED"
    assert payload["packNo"] == "40"
    assert payload["grossWt"] == "296.46"
    assert payload["netWt"] == "218.375"
    assert payload["cusTrafMode"] == "4"
    assert payload["transMode"] == "3"
    assert payload["supvModeCdde"] == "0110"
    assert payload["cusIEFlag"] == "E"
    assert payload["agentCode"] == "4403180867"
    assert payload["agentName"] == "深圳市泰洲物流有限公司"
    assert payload["agentScc"] == ""
    assert payload["iePort"] == "莲塘口岸"
    assert declaration.head["iePort"].status.value == "needs_review"
    assert payload["wrapType"] == "其它"
    assert declaration.head["wrapType"].status.value == "needs_review"
    assert len(payload["tdecGoodsitemsVoArr"]) == 1
    goods = payload["tdecGoodsitemsVoArr"][0]
    assert goods["gno"] == "1"
    assert goods["gname"] == "表壳配件/壳体"
    assert goods["gunit"] == "008"
    assert goods["cusOriginCountry"] == "CHN"
    assert goods["destinationCountry"] == "HKG"
    assert goods["districtCode"] == "44139"
    assert payload["tdecContasVoArr"] == []
    assert payload["_meta"]["source_roles"] == ["draft", "packing"]
    assert "SHOULD-NOT-ASSEMBLE" not in payload.values()


def test_commercial_mismatch_keeps_draft_value() -> None:
    document = _parse({"一般贸易出口": _draft, "装箱单": _mismatch_packing}, "mismatch.xlsx")
    declaration = assemble_declaration(document)
    assert declaration.value_of("packNo") == "40"
    assert declaration.head["packNo"].status.value == "needs_review"
    assert "packNo:head_mismatch" in declaration.review_reasons


def test_net_only_does_not_copy_to_gross() -> None:
    document = _parse({"总箱单": _net_only_packing}, "net-only.xlsx")
    declaration = assemble_declaration(document)
    assert declaration.has_draft is False
    assert declaration.value_of("netWt") == "2825.47"
    assert declaration.value_of("grossWt") in {None, ""}
    assert declaration.head["grossWt"].status.value == "needs_review"
    assert "grossWt:net_is_not_gross" in declaration.review_reasons
    payload = declaration_payload(declaration)
    assert payload["grossWt"] == ""
    assert payload["netWt"] == "2825.47"
    assert payload["supvModeCdde"] == ""
    assert payload["cutMode"] == ""
    assert payload["iePort"] == ""
    assert "supvModeCdde:customs_empty" in declaration.review_reasons


def test_no_draft_fills_commercial_and_leaves_customs_empty() -> None:
    document = _parse(
        {"总箱单": _net_only_packing, "发票": _invoice, "Sheet3": _auxiliary},
        "guoguang.xlsx",
    )
    declaration = assemble_declaration(document)
    payload = declaration_payload(declaration)
    assert declaration.has_draft is False
    assert payload["contrNo"] == "26VN0502"
    assert payload["consignorEname"] == "GUOGUANG ACOUSTICS (VIETNAM) COMPANY LIMITED"
    assert payload["tradeName"] == "GUOGUANG ELECTRIC COMPANY LIMITED."
    assert payload["supvModeCdde"] == ""
    assert payload["cutMode"] == ""
    assert payload["customMaster"] == ""
    assert payload["wrapType"] == ""
    assert payload["agentCode"] == ""
    assert len(payload["tdecGoodsitemsVoArr"]) == 1
    assert payload["tdecGoodsitemsVoArr"][0]["gname"] == "贴纸"
    assert payload["tdecGoodsitemsVoArr"][0]["gqty"] == "100"
    assert "SHOULD-NOT-ASSEMBLE" not in payload.values()
    assert "auxiliary" not in declaration.source_roles


def test_agent_not_taken_from_file() -> None:
    document = _parse({"一般贸易出口": _draft}, "draft-only.xlsx")
    declaration = assemble_declaration(document)
    assert declaration.value_of("agentCode") is None
    assert declaration.value_of("agentName") is None


@pytest.mark.skipif(not REAL_HENGXIN.exists(), reason="本地恒信样本不在 CI")
def test_hengxin_sample_one_declaration() -> None:
    document = parse_excel(
        REAL_HENGXIN.read_bytes(),
        file_id="hx-real",
        filename=REAL_HENGXIN.name,
    )
    payload = declaration_payload(assemble_declaration(document))
    assert payload["contrNo"] == "HDX2026-251"
    assert payload["packNo"] == "40"
    assert payload["grossWt"] == "296.46"
    assert payload["transMode"] == "3"
    assert payload["cusTrafMode"] == "4"
    assert payload["supvModeCdde"] == "0110"
    assert len(payload["tdecGoodsitemsVoArr"]) >= 1
    assert payload["tdecGoodsitemsVoArr"][0]["gname"]


@pytest.mark.skipif(not REAL_GUOGUANG.exists(), reason="本地国光样本不在 CI")
def test_guoguang_sample_one_declaration() -> None:
    document = parse_excel(
        REAL_GUOGUANG.read_bytes(),
        file_id="gg-real",
        filename=REAL_GUOGUANG.name,
    )
    payload = declaration_payload(assemble_declaration(document))
    assert payload["contrNo"] == "26VN0502"
    assert payload["consignorEname"] == "GUOGUANG ACOUSTICS (VIETNAM) COMPANY LIMITED"
    assert payload["supvModeCdde"] == ""
    assert payload["cutMode"] == ""
    assert len(payload["tdecGoodsitemsVoArr"]) >= 1
    assert payload["tdecGoodsitemsVoArr"][0]["gname"]
