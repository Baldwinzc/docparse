from __future__ import annotations

import io
import re
from datetime import datetime
from pathlib import Path

import pytest

pytest.importorskip("openpyxl")

from openpyxl import Workbook
from openpyxl.styles import Border, Side

from docparse.adapters.parsers.excel import parse_excel

_DEMO = Path("/Users/baldwin/Desktop/taizhou/AI识别Demo")
REAL_HENGXIN = _DEMO / "（恒信）一般贸易草单HDX260251BLU.xlsx"
REAL_GUOGUANG = _DEMO / "（国光）箱单发票合同26VN0502-1.xlsx"


def _thin() -> Border:
    line = Side(style="thin")
    return Border(left=line, right=line, top=line, bottom=line)


def _box_workbook() -> bytes:
    book = Workbook()
    draft = book.active
    draft.title = "一般贸易出口"
    packing = book.create_sheet("装箱单")

    packing["G58"] = 40
    packing["I58"] = 296.46
    packing["H58"] = 218.375

    draft.merge_cells("A3:D3")
    draft["A3"] = "境内发货人"
    draft.merge_cells("A4:D4")
    draft["A4"] = "惠州市恒德信精密科技有限公司441394164D"

    draft["E3"] = "出境关别"
    draft["E4"] = "莲塘口岸"

    draft.merge_cells("A5:D5")
    draft["A5"] = "境外收货人"
    draft.merge_cells("A6:D6")
    draft["A6"] = "BLU PRECISION LIMITED"

    draft["E5"] = "运输方式"
    draft["E6"] = "公路运输"

    draft.merge_cells("A7:D7")
    draft["A7"] = "生产销售单位"
    draft.merge_cells("A8:D8")
    draft["A8"] = "惠州市恒德信精密科技有限公司"

    draft["E7"] = "监管方式"
    draft["E8"] = "一般贸易"
    draft.merge_cells("I7:J7")
    draft["I7"] = "征免性质"
    draft["I8"] = "一般征税"

    draft.merge_cells("A9:D9")
    draft["A9"] = "合同协议号"
    draft.merge_cells("A10:D10")
    draft["A10"] = "HDX2026-251"

    draft.merge_cells("A11:C11")
    draft["A11"] = "包装种类"
    draft["D11"] = "件数"
    draft["E11"] = "毛重（千克）"
    draft["F11"] = "净重（千克）"
    draft.merge_cells("G11:H11")
    draft["G11"] = "成交方式"

    draft.merge_cells("A12:C12")
    draft["A12"] = "其它"
    draft["D12"] = "=装箱单!G58"
    draft["E12"] = "=装箱单!I58"
    draft["F12"] = "=装箱单!H58"
    draft.merge_cells("G12:H12")
    draft["G12"] = "FOB"

    headers = [
        "项号",
        "商品编号",
        "商品名称及规格型号",
        "申报要素",
        "数量PCS",
        "计价单位",
        "重量KG",
        "单价",
        "总价",
    ]
    for col, header in enumerate(headers, start=1):
        draft.cell(17, col, header)
    draft["A18"] = 1
    draft["B18"] = "9111900000"
    draft["C18"] = "表壳配件/壳体"
    draft["D18"] = "不锈钢"
    draft["E18"] = 150
    draft["F18"] = "只"
    draft["G18"] = 5.74
    draft["H18"] = 98
    draft["I18"] = "=E18*H18"

    for row in draft.iter_rows(min_row=3, max_row=18, min_col=1, max_col=9):
        for cell in row:
            cell.border = _thin()

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _pairs(sheet) -> dict[str, str]:
    return {item.key: item.value for item in sheet.key_values}


def test_box_form_key_values_and_table() -> None:
    document = parse_excel(_box_workbook(), file_id="f1", filename="draft.xlsx")
    assert {sheet.name for sheet in document.sheets} == {"一般贸易出口", "装箱单"}
    draft = next(sheet for sheet in document.sheets if sheet.name == "一般贸易出口")
    pairs = _pairs(draft)

    assert pairs["境内发货人"] == "惠州市恒德信精密科技有限公司441394164D"
    assert pairs["境外收货人"] == "BLU PRECISION LIMITED"
    assert pairs["合同协议号"] == "HDX2026-251"
    assert pairs["运输方式"] == "公路运输"
    assert pairs["监管方式"] == "一般贸易"
    assert pairs["征免性质"] == "一般征税"
    assert pairs["件数"] == "40"
    assert pairs["毛重（千克）"] == "296.46"
    assert pairs["净重（千克）"] == "218.375"
    assert pairs["成交方式"] == "FOB"

    weight = next(cell for cell in draft.cells if cell.address == "E12")
    assert weight.formula == "=装箱单!I58"
    assert weight.value == "296.46"

    assert draft.tables
    table = draft.tables[0]
    assert "项号" in table.headers
    assert "商品编号" in table.headers
    assert table.rows[0]["项号"] == "1"
    assert table.rows[0]["商品编号"] == "9111900000"
    assert table.rows[0]["商品名称及规格型号"] == "表壳配件/壳体"
    assert table.rows[0]["总价"] == "14700"


def _alias_workbook() -> bytes:
    book = Workbook()
    packing = book.active
    packing.title = "总箱单"
    packing["A1"] = "物料名称"
    packing["B1"] = "出货数量"
    packing["C1"] = "N.W."
    packing["D1"] = "G.W ."
    packing["A2"] = "贴纸"
    packing["B2"] = 100
    packing["C2"] = 0.0013
    packing["D2"] = 0.5013

    draft = book.create_sheet("草单")
    draft["A3"] = "毛重"
    draft["A4"] = "12.5"
    draft["B3"] = "货物存放地点"
    draft["B4"] = "惠州仓库"

    for row in packing.iter_rows(min_row=1, max_row=2, min_col=1, max_col=4):
        for cell in row:
            cell.border = _thin()
    for row in draft.iter_rows(min_row=3, max_row=4, min_col=1, max_col=2):
        for cell in row:
            cell.border = _thin()

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def test_alias_headers_and_box_labels() -> None:
    document = parse_excel(_alias_workbook(), file_id="a1", filename="alias.xlsx")
    packing = next(sheet for sheet in document.sheets if sheet.name == "总箱单")
    assert packing.tables
    table = packing.tables[0]
    assert "物料名称" in table.headers
    assert "出货数量" in table.headers
    assert table.rows[0]["物料名称"] == "贴纸"
    assert table.rows[0]["出货数量"] == "100"

    draft = next(sheet for sheet in document.sheets if sheet.name == "草单")
    pairs = _pairs(draft)
    assert pairs["毛重"] == "12.5"
    assert pairs["货物存放地点"] == "惠州仓库"


def test_box_label_row_is_not_a_table() -> None:
    document = parse_excel(_box_workbook(), file_id="f1", filename="draft.xlsx")
    draft = next(sheet for sheet in document.sheets if sheet.name == "一般贸易出口")
    header_rows = {table.header_row for table in draft.tables}
    assert 11 not in header_rows
    assert 17 in header_rows


def _kv_colon_workbook() -> bytes:
    book = Workbook()
    packing = book.active
    packing.title = "装箱单"
    packing["I4"] = "Invoice No.﹕"
    packing["J4"] = "HDX260251"
    packing["I5"] = "Date：2026-8-6"
    packing["A7"] = "Bill To : BLU PRECISION LIMITED"

    invoice = book.create_sheet("发票")
    invoice["A6"] = "日期DATE:"
    invoice["B6"] = datetime(2026, 3, 15, 10, 30, 0)
    invoice["F6"] = "发票号INVOICE NO.:"
    invoice["G6"] = "26VN0502-1"
    invoice["A11"] = "运输工具 SHIPPED PER:"
    invoice["B11"] = "BY TRUCK"

    for row in packing.iter_rows(min_row=4, max_row=7, min_col=1, max_col=10):
        for cell in row:
            cell.border = _thin()
    for row in invoice.iter_rows(min_row=6, max_row=11, min_col=1, max_col=7):
        for cell in row:
            cell.border = _thin()

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _dual_header_workbook() -> bytes:
    book = Workbook()
    contract = book.active
    contract.title = "合同"
    contract["A14"] = "(1)货物名称"
    contract["C14"] = "(3)数量"
    contract["E14"] = "(5)单价"
    contract["F14"] = "(6)总值"
    contract["A15"] = "Description"
    contract["C15"] = "Quantity"
    contract["E15"] = "Unit Price(HKD)"
    contract["F15"] = "Total Amount(HKD)"
    contract["A16"] = 1
    contract["C16"] = 150
    contract["E16"] = 98
    contract["F16"] = 14700
    contract["A17"] = 2
    contract["C17"] = 73
    contract["E17"] = 1398
    contract["F17"] = 10526.94

    for row in contract.iter_rows(min_row=14, max_row=17, min_col=1, max_col=6):
        for cell in row:
            cell.border = _thin()

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def test_special_colon_and_datetime_stay_in_key_values() -> None:
    document = parse_excel(_kv_colon_workbook(), file_id="k1", filename="kv.xlsx")
    packing = next(sheet for sheet in document.sheets if sheet.name == "装箱单")
    pairs = _pairs(packing)
    assert pairs["Invoice No."] == "HDX260251"
    assert pairs["Date"] == "2026-8-6"
    assert pairs["Bill To"] == "BLU PRECISION LIMITED"

    invoice = next(sheet for sheet in document.sheets if sheet.name == "发票")
    pairs = _pairs(invoice)
    assert pairs["日期DATE"] == "2026-03-15 10:30:00"
    assert pairs["发票号INVOICE NO."] == "26VN0502-1"
    assert pairs["运输工具 SHIPPED PER"] == "BY TRUCK"
    assert "2026-03-15 10" not in pairs
    assert all("10:30:00" not in item.key for item in invoice.key_values)


def test_dual_header_english_row_is_not_goods() -> None:
    document = parse_excel(_dual_header_workbook(), file_id="d1", filename="contract.xlsx")
    contract = document.sheets[0]
    assert contract.tables
    table = contract.tables[0]
    assert table.header_row == 14
    assert table.header_rows == [14, 15]
    assert any("货物名称" in header for header in table.headers)
    assert any("Description" in header for header in table.headers)
    assert table.rows
    first = table.rows[0]
    assert "Description" not in first.values()
    assert "1" in first.values()
    assert "150" in first.values()


def test_hengxin_draft_goods_table_still_single_header() -> None:
    document = parse_excel(_box_workbook(), file_id="f1", filename="draft.xlsx")
    draft = next(sheet for sheet in document.sheets if sheet.name == "一般贸易出口")
    table = draft.tables[0]
    assert table.header_row == 17
    assert table.header_rows == [17]
    assert table.rows[0]["项号"] == "1"
    assert table.rows[0]["商品编号"] == "9111900000"


@pytest.mark.skipif(not REAL_HENGXIN.exists(), reason="本地恒信样本不在 CI")
def test_hengxin_sample_keeps_formulas_and_goods_table() -> None:
    document = parse_excel(
        REAL_HENGXIN.read_bytes(),
        file_id="hx",
        filename=REAL_HENGXIN.name,
    )
    assert len(document.sheets) >= 4
    draft = next(sheet for sheet in document.sheets if "一般贸易" in sheet.name)
    pairs = _pairs(draft)
    assert pairs["合同协议号"] == "HDX2026-251"
    assert pairs["境外收货人"] == "BLU PRECISION LIMITED"
    assert pairs["出境关别"] == "莲塘口岸"
    assert pairs["离境口岸"] == "莲塘口岸"
    assert pairs["件数"] == "40"
    assert pairs["毛重（千克）"] == "296.46"
    assert draft.tables
    first = draft.tables[0].rows[0]
    assert first["项号"] == "1"
    assert first["商品编号"] == "9111900000"


@pytest.mark.skipif(not REAL_HENGXIN.exists(), reason="本地恒信样本不在 CI")
def test_hengxin_packing_and_contract_layout() -> None:
    document = parse_excel(
        REAL_HENGXIN.read_bytes(),
        file_id="hx2",
        filename=REAL_HENGXIN.name,
    )
    packing = next(sheet for sheet in document.sheets if "箱" in sheet.name)
    pairs = _pairs(packing)
    assert pairs.get("Invoice No.") == "HDX260251"
    contract = next(sheet for sheet in document.sheets if sheet.name == "合同")
    assert contract.tables
    table = contract.tables[0]
    assert table.rows
    assert "Description" not in table.rows[0].values()


@pytest.mark.skipif(not REAL_GUOGUANG.exists(), reason="本地国光样本不在 CI")
def test_guoguang_invoice_number_and_date_in_kv() -> None:
    document = parse_excel(
        REAL_GUOGUANG.read_bytes(),
        file_id="gg",
        filename=REAL_GUOGUANG.name,
    )
    packing = next(sheet for sheet in document.sheets if "箱" in sheet.name)
    pairs = _pairs(packing)
    date_keys = [key for key in pairs if "DATE" in key.upper() or key.startswith("日期")]
    assert date_keys, pairs
    date_value = pairs[date_keys[0]]
    assert "2026" in date_value or date_value.replace(".", "").isdigit()
    invoice_sheet = next(sheet for sheet in document.sheets if "发票" in sheet.name)
    invoice_pairs = _pairs(invoice_sheet)
    invoice_keys = [key for key in invoice_pairs if "INVOICE" in key.upper() or "发票" in key]
    assert invoice_keys, invoice_pairs
    assert any("26VN0502" in invoice_pairs[key] for key in invoice_keys)


def _value_shape_workbook() -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = "值域"

    sheet["A1"] = "出口日期:备注栏"
    sheet["A2"] = "2026-07-17 00:00:00"

    sheet["B1"] = "件数"
    sheet["C1"] = "惠州市恒德信精密科技有限公司"
    sheet["B2"] = 40

    sheet["D1"] = "货物存放地点:惠州仓库"

    sheet["E1"] = "Invoice No.﹕不像单号"
    sheet["F1"] = "隔壁不是单号"

    for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
        for cell in row:
            cell.border = _thin()

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _unconstrained_key_items(sheet) -> list:
    return [item for item in sheet.key_values if item.key == "货物存放地点"]


def test_value_shape_prefers_datetime_below_over_same_cell_fragment() -> None:
    document = parse_excel(_value_shape_workbook(), file_id="v1", filename="shape.xlsx")
    sheet = document.sheets[0]
    dates = [item for item in sheet.key_values if item.key == "出口日期"]
    assert len(dates) == 1
    assert dates[0].strategy == "below"
    assert dates[0].value == "2026-07-17 00:00:00"


def test_value_shape_prefers_numeric_pack_no_over_company_name() -> None:
    document = parse_excel(_value_shape_workbook(), file_id="v1", filename="shape.xlsx")
    sheet = document.sheets[0]
    packs = [item for item in sheet.key_values if item.key == "件数"]
    assert len(packs) == 1
    assert packs[0].value == "40"
    assert packs[0].strategy == "below"


def test_unconstrained_key_keeps_same_cell_like_issue_15() -> None:
    document = parse_excel(_value_shape_workbook(), file_id="v1", filename="shape.xlsx")
    sheet = document.sheets[0]
    items = _unconstrained_key_items(sheet)
    assert len(items) == 1
    assert items[0].strategy == "same_cell"
    assert items[0].value == "惠州仓库"


def test_unlike_values_drop_the_key() -> None:
    document = parse_excel(_value_shape_workbook(), file_id="v1", filename="shape.xlsx")
    sheet = document.sheets[0]
    assert "Invoice No." not in _pairs(sheet)


_SAMPLES = Path("/Users/baldwin/Desktop/taizhou/补充测试")
REAL_GSC = _SAMPLES / "GSC出仓QPGSCO260617006A.xlsx"
REAL_GSRUA = _SAMPLES / "GSRUA26601CLLG01(1).xlsx"

_PLACEHOLDER_VALUE = re.compile(r"^[\(（]\s*[\)）]")


def _flat_gap_workbook() -> bytes:
    """GSC 平表：标签与值隔空列、below 是下一行的标签。"""
    book = Workbook()
    sheet = book.active
    sheet.title = "出境备案清单"

    sheet["A2"] = "出口口岸"
    sheet["C2"] = "5304"
    sheet["A3"] = "经营单位"
    sheet["C3"] = "440356K004"
    sheet["F2"] = "备案号"
    sheet["G2"] = "T5339W000208"
    sheet["F3"] = "运输工具"
    sheet["A6"] = "许可证号"
    sheet["A7"] = "批准文号"

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def test_right_skips_blank_columns_and_rejects_label_below() -> None:
    document = parse_excel(_flat_gap_workbook(), file_id="g1", filename="flat.xlsx")
    sheet = document.sheets[0]
    pairs = _pairs(sheet)
    assert pairs["出口口岸"] == "5304"
    assert pairs["备案号"] == "T5339W000208"
    assert pairs["经营单位"] == "440356K004"
    values = {item.value for item in sheet.key_values}
    assert "经营单位" not in values
    assert "运输工具" not in values
    assert "批准文号" not in values
    assert "许可证号" not in pairs


def _fake_header_workbook() -> bytes:
    """GSC R6/R8：长套话凑 token 的数据行不当表头。"""
    book = Workbook()
    sheet = book.active
    sheet.title = "出境备案清单"

    sheet["D6"] = "运抵国"
    sheet["E6"] = "中国台湾"
    sheet["F6"] = "指运港"
    sheet["G6"] = "中国台湾"
    sheet["I6"] = "境内货源地"
    sheet["J6"] = "44036"
    sheet["T6"] = "关联收发货单位海关编码"

    sheet["D7"] = "成交方式"
    sheet["E7"] = "FOB"
    sheet["D8"] = "件数"
    sheet["E8"] = "230.000000"
    sheet["F8"] = "包装种类"
    sheet["G8"] = "纸制或纤维板制盒/箱"
    sheet["I8"] = "毛重"
    sheet["J8"] = "1560"
    sheet["K8"] = "净重"
    sheet["L8"] = "785.02760"

    headers = ["项号", "海关编码", "商品名称", "规格型号", "数量"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(11, col, header)
    sheet["A12"] = "1"
    sheet["B12"] = "9503002900"
    sheet["C12"] = "塑胶动漫造形"
    sheet["D12"] = "MAX FACTORY 品牌"
    sheet["E12"] = "96"

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def test_data_rows_with_numbers_are_not_fake_headers() -> None:
    document = parse_excel(_fake_header_workbook(), file_id="g2", filename="fake.xlsx")
    sheet = document.sheets[0]
    header_rows = {table.header_row for table in sheet.tables}
    assert 6 not in header_rows
    assert 8 not in header_rows

    pairs = _pairs(sheet)
    assert pairs["运抵国"] == "中国台湾"
    assert pairs["指运港"] == "中国台湾"
    assert pairs["成交方式"] == "FOB"
    assert float(pairs["件数"]) == 230
    assert pairs["包装种类"] == "纸制或纤维板制盒/箱"
    assert float(pairs["毛重"]) == 1560
    assert float(pairs["净重"]) == 785.0276

    goods = next(table for table in sheet.tables if table.header_row == 11)
    assert goods.rows[0]["海关编码"] == "9503002900"


def _placeholder_workbook() -> bytes:
    """GSRUA：占位格 (  ) 不当值、不当表头，框表标签行带占位格仍成立。"""
    book = Workbook()
    sheet = book.active
    sheet.title = "报关单"

    sheet["A2"] = "预录入编号："
    sheet["G2"] = "申报口岸:"
    sheet["G3"] = "(    )"
    sheet["F3"] = "出境关别"
    sheet["H3"] = "出口日期"
    sheet["F7"] = "监管方式"
    sheet["G7"] = "(    )"
    sheet["H7"] = "征免性质"
    sheet["I7"] = "(    )"

    sheet["A11"] = "包装类型"
    sheet["C11"] = "(    )"
    sheet["F11"] = "件数"
    sheet["G11"] = "毛重（千克）"
    sheet["H11"] = "净重（千克）"
    sheet["J11"] = "成交方式"
    sheet["K11"] = "(  )"
    sheet["L11"] = "运费"
    sheet["A12"] = "胶合板箱,纸箱"
    sheet["F12"] = "10"
    sheet["G12"] = "1725"
    sheet["H12"] = "1017.72"
    sheet["J12"] = "DAT Selyatino"

    headers = ["项号", "商品编号", "商品名称"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(15, col, header)
    sheet["A16"] = "1"
    sheet["B16"] = "9503002900"
    sheet["C16"] = "塑胶动漫造形"
    sheet["A17"] = "(    )"
    sheet["B17"] = "(    )"
    sheet["C17"] = "(    )"
    sheet["A18"] = "2"
    sheet["B18"] = "9503002901"
    sheet["C18"] = "第二行"

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def test_placeholder_cells_are_blank_not_values() -> None:
    document = parse_excel(_placeholder_workbook(), file_id="p1", filename="ph.xlsx")
    sheet = document.sheets[0]
    pairs = _pairs(sheet)
    assert "申报口岸" not in pairs
    assert "出境关别" not in pairs
    assert "征免性质" not in pairs
    assert all(_PLACEHOLDER_VALUE.match(value) is None for value in pairs.values())

    header_rows = {table.header_row for table in sheet.tables}
    assert 11 not in header_rows
    assert pairs["包装类型"] == "胶合板箱,纸箱"
    assert pairs["件数"] == "10"
    assert float(pairs["毛重（千克）"]) == 1725
    assert float(pairs["净重（千克）"]) == 1017.72
    assert pairs["成交方式"] == "DAT Selyatino"

    goods = next(table for table in sheet.tables if table.header_row == 15)
    assert [row["项号"] for row in goods.rows] == ["1"]


@pytest.mark.skipif(not REAL_GSC.exists(), reason="本地 GSC 样本不在 CI")
def test_real_gsc_flat_layout_kvs_and_goods_table() -> None:
    document = parse_excel(
        REAL_GSC.read_bytes(),
        file_id="gsc",
        filename=REAL_GSC.name,
    )
    draft = document.sheets[0]
    pairs = _pairs(draft)
    assert pairs["出口口岸"] == "5304"
    assert pairs["备案号"] == "T5339W000208"
    assert pairs["贸易方式"] == "区内物流货物"
    assert pairs["运抵国"] == "中国台湾"
    assert pairs["指运港"] == "中国台湾"
    assert float(pairs["件数"]) == 230
    assert pairs["包装种类"] == "纸制或纤维板制盒/箱"
    assert float(pairs["毛重"]) == 1560
    assert float(pairs["净重"]) == 785.0276
    assert pairs["成交方式"] == "FOB"
    assert pairs["贸易国"] == "中国台湾"

    values = {item.value for item in draft.key_values}
    assert "经营单位" not in values
    assert "运输工具" not in values
    assert "批准文号" not in values

    header_rows = {table.header_row for table in draft.tables}
    assert 6 not in header_rows
    assert 8 not in header_rows
    goods = next(table for table in draft.tables if table.header_row == 11)
    assert len(goods.rows) == 50
    assert goods.rows[0]["海关编码"] == "9503002900"


@pytest.mark.skipif(not REAL_GSRUA.exists(), reason="本地 GSRUA 样本不在 CI")
def test_real_gsrua_box_row_with_placeholders() -> None:
    document = parse_excel(
        REAL_GSRUA.read_bytes(),
        file_id="gsrua",
        filename=REAL_GSRUA.name,
    )
    draft = document.sheets[0]
    pairs = _pairs(draft)
    assert pairs["包装类型"] == "胶合板箱,纸箱"
    assert pairs["件数"] == "10"
    assert float(pairs["毛重（千克）"]) == 1725
    assert float(pairs["净重（千克）"]) == 1017.72
    assert pairs["成交方式"] == "DAT Selyatino"
    assert "出境关别" not in pairs
    assert "征免性质" not in pairs
    assert "离境口岸" not in pairs
    assert "申报口岸" not in pairs
    assert all(_PLACEHOLDER_VALUE.match(value) is None for value in pairs.values())

    header_rows = {table.header_row for table in draft.tables}
    assert 11 not in header_rows
    goods = next(table for table in draft.tables if table.header_row == 19)
    assert len(goods.rows) == 101
