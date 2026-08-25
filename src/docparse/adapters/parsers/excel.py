import csv
import io
import re
from datetime import date, datetime
from uuid import uuid4

from docparse.adapters.parsers.layout import split_sheet
from docparse.domain.ir import Cell, CellBorder, DocumentIR, Sheet
from docparse.extraction.sheet_role import classify_sheets

_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_XLS = "application/vnd.ms-excel"
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0"
_SHEET_REF = re.compile(
    r"^=(?:'([^']+)'|([^'!]+))!(\$?[A-Z]{1,3}\$?\d+)$",
    re.IGNORECASE,
)
_MUL_REF = re.compile(
    r"^=(\$?[A-Z]{1,3}\$?\d+)\s*\*\s*(\$?[A-Z]{1,3}\$?\d+)$",
    re.IGNORECASE,
)


def parse_excel(data: bytes, *, file_id: str, filename: str) -> DocumentIR:
    if filename.lower().endswith(".csv"):
        return _parse_csv(data, file_id=file_id, filename=filename)
    if data.startswith(_OLE2_MAGIC):
        return _parse_xls(data, file_id=file_id, filename=filename)
    return _parse_xlsx(data, file_id=file_id, filename=filename)


def _parse_xls(data: bytes, *, file_id: str, filename: str) -> DocumentIR:
    try:
        import xlrd
    except ImportError:
        return DocumentIR(
            document_id=uuid4().hex,
            file_id=file_id,
            filename=filename,
            media_type=_XLS,
            warnings=["未安装 xlrd，.xls 仅登记未解析。pip install 'docparse[excel]'"],
        )

    book = xlrd.open_workbook(file_contents=data, formatting_info=True)
    parsed = [_read_xls_sheet(book, sheet) for sheet in book.sheets()]
    return _finish_document(parsed, file_id=file_id, filename=filename, media_type=_XLS)


def _read_xls_sheet(book, raw_sheet) -> Sheet:
    merges = _xls_merge_origins(raw_sheet)
    covered = _xls_covered_cells(raw_sheet)
    cells: list[Cell] = []
    for row_idx in range(raw_sheet.nrows):
        for col_idx in range(raw_sheet.ncols):
            address = f"{_col_letter(col_idx + 1)}{row_idx + 1}"
            if address in covered and address not in merges:
                continue
            item = raw_sheet.cell(row_idx, col_idx)
            display = _xls_value(book, item)
            if not display and address not in merges:
                continue
            cells.append(
                Cell(
                    address=address,
                    value=display,
                    raw_value=display or None,
                    row=row_idx + 1,
                    column=col_idx + 1,
                    merge_range=merges.get(address),
                    border=None,
                )
            )
    return Sheet(name=raw_sheet.name, cells=cells)


def _xls_value(book, item) -> str:
    import xlrd

    if item.ctype == xlrd.XL_CELL_ERROR:
        return ""
    if item.ctype == xlrd.XL_CELL_DATE:
        try:
            return _stringify(xlrd.xldate_as_datetime(item.value, book.datemode))
        except (ValueError, OverflowError):
            return _stringify(item.value)
    return _stringify(item.value)


def _xls_merge_origins(sheet) -> dict[str, str]:
    origins: dict[str, str] = {}
    for rlo, rhi, clo, chi in sheet.merged_cells:
        start = f"{_col_letter(clo + 1)}{rlo + 1}"
        end = f"{_col_letter(chi)}{rhi}"
        origins[start] = f"{start}:{end}"
    return origins


def _xls_covered_cells(sheet) -> set[str]:
    covered: set[str] = set()
    for rlo, rhi, clo, chi in sheet.merged_cells:
        for row in range(rlo, rhi):
            for col in range(clo, chi):
                covered.add(f"{_col_letter(col + 1)}{row + 1}")
    return covered


def _parse_xlsx(data: bytes, *, file_id: str, filename: str) -> DocumentIR:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return DocumentIR(
            document_id=uuid4().hex,
            file_id=file_id,
            filename=filename,
            media_type=_XLSX,
            warnings=["未安装 openpyxl，Excel 仅登记未解析。pip install 'docparse[excel]'"],
        )

    raw_book = load_workbook(io.BytesIO(data), data_only=False)
    value_book = load_workbook(io.BytesIO(data), data_only=True)
    value_sheets = {sheet.title: sheet for sheet in value_book.worksheets}

    parsed: list[Sheet] = []
    grids: dict[str, dict[str, str]] = {}
    pending: list[tuple[str, Cell]] = []

    for raw_sheet in raw_book.worksheets:
        value_sheet = value_sheets.get(raw_sheet.title)
        sheet, grid, formulas = _read_sheet(raw_sheet, value_sheet)
        grids[sheet.name] = grid
        parsed.append(sheet)
        pending.extend((sheet.name, cell) for cell in formulas)

    _resolve_formulas(grids, pending)
    return _finish_document(parsed, file_id=file_id, filename=filename, media_type=_XLSX)


def _finish_document(
    parsed: list[Sheet],
    *,
    file_id: str,
    filename: str,
    media_type: str,
) -> DocumentIR:
    sheets = [split_sheet(sheet) for sheet in parsed]
    document = DocumentIR(
        document_id=uuid4().hex,
        file_id=file_id,
        filename=filename,
        media_type=media_type,
        sheets=sheets,
        raw_text="",
    )
    classify_sheets(document)
    document.raw_text = "\n".join(
        f"{sheet.name}!{cell.address}:{cell.value}"
        for sheet in document.sheets
        for cell in sheet.cells
    )
    return document


def _read_sheet(raw_sheet, value_sheet) -> tuple[Sheet, dict[str, str], list[Cell]]:
    merges = _merge_origins(raw_sheet)
    covered = _covered_cells(raw_sheet)
    cells: list[Cell] = []
    grid: dict[str, str] = {}
    formulas: list[Cell] = []

    for row in raw_sheet.iter_rows():
        for item in row:
            address = item.coordinate
            if address in covered and address not in merges:
                continue
            cached = None
            if value_sheet is not None:
                cached = value_sheet[address].value
            raw = item.value
            formula = raw if isinstance(raw, str) and raw.startswith("=") else None
            fallback = None if formula else raw
            display = _stringify(cached if cached is not None else fallback)
            if not display and not formula and address not in merges:
                continue
            cell = Cell(
                address=address,
                value=display,
                raw_value=formula or _stringify(item.value) or None,
                row=item.row,
                column=item.column,
                merge_range=merges.get(address),
                formula=formula,
                border=_border(item),
            )
            cells.append(cell)
            if display:
                grid[address] = display
            if formula and not display:
                formulas.append(cell)

    return Sheet(name=raw_sheet.title, cells=cells), grid, formulas


def _merge_origins(sheet) -> dict[str, str]:
    origins: dict[str, str] = {}
    for item in sheet.merged_cells.ranges:
        origins[item.start_cell.coordinate] = str(item)
    return origins


def _covered_cells(sheet) -> set[str]:
    covered: set[str] = set()
    for item in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = item.bounds
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                covered.add(f"{_col_letter(col)}{row}")
    return covered


def _col_letter(index: int) -> str:
    letters = ""
    while index:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _border(item) -> CellBorder | None:
    border = item.border
    if border is None:
        return None
    sides = CellBorder(
        left=bool(border.left and border.left.style),
        right=bool(border.right and border.right.style),
        top=bool(border.top and border.top.style),
        bottom=bool(border.bottom and border.bottom.style),
    )
    if not any((sides.left, sides.right, sides.top, sides.bottom)):
        return None
    return sides


def _stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".10g")
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _resolve_formulas(
    grids: dict[str, dict[str, str]],
    pending: list[tuple[str, Cell]],
) -> None:
    for _ in range(8):
        leftover: list[tuple[str, Cell]] = []
        progress = False
        for sheet_name, cell in pending:
            resolved = _eval_formula(cell.formula or "", sheet_name, grids)
            if resolved is None:
                leftover.append((sheet_name, cell))
                continue
            cell.value = resolved
            grids.setdefault(sheet_name, {})[cell.address] = resolved
            progress = True
        pending = leftover
        if not progress or not pending:
            break


def _eval_formula(formula: str, sheet_name: str, grids: dict[str, dict[str, str]]) -> str | None:
    match = _SHEET_REF.match(formula.replace("$", ""))
    if match:
        target = match.group(1) or match.group(2)
        address = match.group(3).upper()
        value = grids.get(target, {}).get(address)
        return value or None
    match = _MUL_REF.match(formula.replace("$", ""))
    if match:
        left = _as_number(grids.get(sheet_name, {}).get(match.group(1).upper()))
        right = _as_number(grids.get(sheet_name, {}).get(match.group(2).upper()))
        if left is None or right is None:
            return None
        return _stringify(left * right)
    return None


def _as_number(text: str | None) -> float | None:
    if text is None or text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_csv(data: bytes, *, file_id: str, filename: str) -> DocumentIR:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    cells: list[Cell] = []
    parts: list[str] = []
    for r_idx, row in enumerate(reader, start=1):
        for c_idx, value in enumerate(row, start=1):
            cleaned = value.strip()
            if not cleaned:
                continue
            address = f"R{r_idx}C{c_idx}"
            cells.append(Cell(address=address, value=cleaned, row=r_idx, column=c_idx))
            parts.append(f"{address}:{cleaned}")
    sheet = split_sheet(Sheet(name="Sheet1", cells=cells))
    document = DocumentIR(
        document_id=uuid4().hex,
        file_id=file_id,
        filename=filename,
        media_type="text/csv",
        sheets=[sheet],
        raw_text="\n".join(parts),
    )
    return classify_sheets(document)
